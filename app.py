import sqlite3
import json
import os
import re
import logging
import secrets
from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from parser import QuoteParser, DellExcelParser
from component_registry import ComponentRegistry
from datetime import datetime

# Data directory — outside the app root by default.
# Override with DATA_DIR env var for custom deployments.
_app_dir = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(_app_dir, '..', 'data'))
DATA_DIR = os.path.abspath(DATA_DIR)
os.makedirs(os.path.join(DATA_DIR, 'uploads'), exist_ok=True)
os.makedirs(os.path.join(DATA_DIR, 'quickspecs'), exist_ok=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(DATA_DIR, 'app.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['DATABASE'] = os.path.join(DATA_DIR, 'quotes.db')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
app.config['SESSION_PERMANENT'] = True

_base = os.environ.get('BASE_URL', '/quotes').rstrip('/')
BASE_HREF = _base + '/'

# Cache-bust static assets on every restart so browsers pick up CSS/JS changes
import time as _time
STATIC_VERSION = str(int(_time.time()))


_DATE_FORMATS = ('%Y-%m-%d', '%m/%d/%Y', '%B %d, %Y', '%B %d, %Y', '%-m/%-d/%Y', '%d-%b-%Y')

def normalize_date(value):
    """Parse any known date string and return YYYY-MM-DD, or None if unparseable."""
    if not value:
        return None
    v = str(value).strip()
    if not v:
        return None
    from datetime import datetime
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return v  # Return as-is if no format matched rather than silently drop


@app.template_filter('is_expired')
def is_expired_filter(expiry_date):
    if not expiry_date:
        return False
    from datetime import date, datetime
    normed = normalize_date(expiry_date)
    try:
        return datetime.strptime(normed, '%Y-%m-%d').date() < date.today()
    except (ValueError, TypeError):
        return False


def _git_branch():
    try:
        import subprocess
        return subprocess.check_output(
            ['git', '-C', _app_dir, 'rev-parse', '--abbrev-ref', 'HEAD'],
            stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        return 'unknown'

GIT_BRANCH = _git_branch()


@app.context_processor
def inject_base_href():
    return {
        'base_href': BASE_HREF,
        'git_branch': GIT_BRANCH,
        'static_version': STATIC_VERSION,
        'current_user': session.get('user_name'),
    }


def _current_user():
    """Return the logged-in username, or 'system' if outside request context."""
    try:
        return session.get('user_name', 'system')
    except RuntimeError:
        return 'system'


def login_required(f):
    """Decorator that redirects to /login when no active session exists."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_name'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_pdf(filepath, max_size_mb=16):
    """
    Validate PDF file for basic security checks.

    Args:
        filepath: Path to PDF file
        max_size_mb: Maximum allowed file size in MB

    Returns:
        Tuple of (is_valid, error_message)
    """
    try:
        # Check file size
        file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        if file_size_mb > max_size_mb:
            return False, f"File size ({file_size_mb:.1f}MB) exceeds maximum allowed ({max_size_mb}MB)"

        # Try to open with pdfplumber to validate structure
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            # Check if PDF has pages
            if len(pdf.pages) == 0:
                return False, "PDF file contains no pages"

            # Check for excessive pages (potential DoS)
            if len(pdf.pages) > 1000:
                return False, "PDF file has too many pages (max 1000)"

            # Try to extract text from first page to verify it's a real PDF
            first_page = pdf.pages[0]
            text = first_page.extract_text()

            # If completely empty, might be suspicious
            if not text or len(text.strip()) == 0:
                # Check if it's an image-based PDF
                if not first_page.images:
                    return False, "PDF appears to be empty or corrupted"

        return True, None

    except Exception as e:
        return False, f"PDF validation failed: {str(e)}"


@app.template_filter('calculate_memory_total')
def calculate_memory_total(memory_items):
    """Calculate total memory in GB from memory items list.

    Args:
        memory_items: List of memory item dicts with 'specs' containing 'capacity_gb'

    Returns:
        Total memory in GB, or 0 if no valid items
    """
    total_gb = 0
    for item in memory_items:
        try:
            quantity = item.get('quantity', 1)
            capacity = item.get('specs', {}).get('capacity_gb', 0)
            if capacity:
                total_gb += quantity * capacity
        except (TypeError, AttributeError, ValueError):
            # Skip items with malformed data
            continue
    return total_gb


def get_db():
    """Get database connection."""
    db = sqlite3.connect(app.config['DATABASE'])
    db.row_factory = sqlite3.Row
    return db


def init_db():
    """Initialize database schema."""
    db = get_db()
    with open(os.path.join(_app_dir, 'schema_normalized.sql'), 'r') as f:
        db.executescript(f.read())
    db.commit()
    db.close()


def migrate_db():
    """Apply incremental schema migrations. Each block is idempotent."""
    db = get_db()
    try:
        db.execute("PRAGMA foreign_keys = OFF")

        tables = lambda: {r[0] for r in db.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        cols   = lambda t: {r[1] for r in db.execute(f"PRAGMA table_info({t})")}

        # ── M01: projects.status CHECK → include 'archived' ──────────────────
        row = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='projects'").fetchone()
        if row and "'archived'" not in row['sql']:
            existing = [c for c in ['id','name','tenant_id','description','comments',
                                    'created_at','updated_at','status'] if c in cols('projects')]
            cs = ', '.join(existing)
            db.execute("""CREATE TABLE projects_migration (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
                tenant_id INTEGER NOT NULL, description TEXT, comments TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'active' CHECK(status IN ('active','inactive','archived')),
                FOREIGN KEY (tenant_id) REFERENCES tenants(id) ON DELETE CASCADE,
                UNIQUE(tenant_id, name))""")
            db.execute(f"INSERT INTO projects_migration ({cs}) SELECT {cs} FROM projects")
            db.execute("DROP TABLE projects")
            db.execute("ALTER TABLE projects_migration RENAME TO projects")
            db.execute("CREATE INDEX IF NOT EXISTS idx_projects_tenant ON projects(tenant_id)")
            db.commit()
            logger.info("M01: projects.status updated")

        # ── M02: tenants.status CHECK → include 'archived' ───────────────────
        row = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='tenants'").fetchone()
        if row and "'archived'" not in row['sql']:
            existing = [c for c in ['id','name','contact_name','created_at','updated_at','status']
                        if c in cols('tenants')]
            cs = ', '.join(existing)
            db.execute("""CREATE TABLE tenants_migration (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE,
                contact_name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'active' CHECK(status IN ('active','inactive','archived')))""")
            db.execute(f"INSERT INTO tenants_migration ({cs}) SELECT {cs} FROM tenants")
            db.execute("DROP TABLE tenants")
            db.execute("ALTER TABLE tenants_migration RENAME TO tenants")
            db.execute("CREATE INDEX IF NOT EXISTS idx_tenant_name ON tenants(name)")
            db.commit()
            logger.info("M02: tenants.status updated")

        # ── M03: quotes.po_comments ───────────────────────────────────────────
        if 'po_comments' not in cols('quotes'):
            db.execute("ALTER TABLE quotes ADD COLUMN po_comments TEXT CHECK(length(po_comments)<=255)")
            db.commit()
            logger.info("M03: quotes.po_comments added")

        # ── M04: manufacturers table ──────────────────────────────────────────
        if 'manufacturers' not in tables():
            db.execute("""CREATE TABLE manufacturers (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE)""")
            db.commit()
            logger.info("M04: manufacturers table created")
        for m in ('Intel','HPE','Dell','Broadcom','Cisco','Samsung',
                  'Seagate','Western Digital','Micron','NVIDIA','AMD'):
            db.execute("INSERT OR IGNORE INTO manufacturers (name) VALUES (?)", (m,))
        db.commit()

        # ── M05: vendors table ────────────────────────────────────────────────
        if 'vendors' not in tables():
            db.execute("""CREATE TABLE vendors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE, code TEXT NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            db.commit()
            logger.info("M05: vendors table created")
        for v in (('HPE','HPE'),('Dell','Dell'),('Cisco','Cisco')):
            db.execute("INSERT OR IGNORE INTO vendors (name,code) VALUES (?,?)", v)
        db.commit()

        # ── M06: quotes.vendor_id FK ──────────────────────────────────────────
        if 'vendor_id' not in cols('quotes'):
            db.execute("ALTER TABLE quotes ADD COLUMN vendor_id INTEGER REFERENCES vendors(id)")
            db.execute("UPDATE quotes SET vendor_id=(SELECT id FROM vendors WHERE code=quotes.vendor)")
            db.commit()
            logger.info("M06: quotes.vendor_id added and populated")

        # ── M07: rename components → learned_components ───────────────────────
        if 'components' in tables() and 'learned_components' not in tables():
            db.execute("ALTER TABLE components RENAME TO learned_components")
            db.commit()
            logger.info("M07: components renamed to learned_components")

        # ── M08: learned_components.manufacturer_id FK (soft) ────────────────
        if 'learned_components' in tables() and 'manufacturer_id' not in cols('learned_components'):
            db.execute("ALTER TABLE learned_components ADD COLUMN manufacturer_id INTEGER REFERENCES manufacturers(id)")
            db.execute("""UPDATE learned_components SET manufacturer_id=
                (SELECT id FROM manufacturers WHERE name=learned_components.manufacturer)""")
            db.commit()
            logger.info("M08: learned_components.manufacturer_id added")

        # ── M09: base_configs table ───────────────────────────────────────────
        if 'base_configs' not in tables():
            db.execute("""CREATE TABLE base_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                config_name TEXT NOT NULL, project_id INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_base_configs_project ON base_configs(project_id)")
            db.commit()
            logger.info("M09: base_configs table created")

        # ── M10: defined_components table ─────────────────────────────────────
        # Only create if bcc does not yet exist — once bcc is present (whether the old
        # defined_components-backed version from M11, or the component_catalog-backed
        # version created by schema_normalized.sql / M16), defined_components is no
        # longer needed and must NOT be re-created (it would re-trigger M16 and wipe bcc).
        if 'defined_components' not in tables() and 'base_config_components' not in tables():
            db.execute("""CREATE TABLE defined_components (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                component_type TEXT, manufacturer_id INTEGER REFERENCES manufacturers(id),
                part_number TEXT, model TEXT, specs TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_defined_comp_type ON defined_components(component_type)")
            db.commit()
            logger.info("M10: defined_components table created")

        # ── M11: base_config_components junction ──────────────────────────────
        if 'base_config_components' not in tables():
            db.execute("""CREATE TABLE base_config_components (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                config_id    INTEGER NOT NULL REFERENCES base_configs(id)    ON DELETE CASCADE,
                component_id INTEGER NOT NULL REFERENCES defined_components(id),
                quantity     INTEGER NOT NULL DEFAULT 1)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_bcc_config    ON base_config_components(config_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_bcc_component ON base_config_components(component_id)")
            db.commit()
            logger.info("M11: base_config_components junction table created")

        # ── M12: drop legacy entered_components ───────────────────────────────
        if 'entered_components' in tables():
            db.execute("DROP TABLE entered_components")
            db.commit()
            logger.info("M12: entered_components dropped (superseded by defined_components)")

        # ── M13a: transaction_types table + seed ──────────────────────────────
        if 'transaction_types' not in tables():
            db.execute("""CREATE TABLE transaction_types (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                code       TEXT NOT NULL UNIQUE,
                label      TEXT NOT NULL,
                description TEXT,
                token_cost INTEGER NOT NULL DEFAULT 1)""")
            db.commit()
            logger.info("M13a: transaction_types table created")

        _TRANSACTION_TYPES = [
            ('add_quote',            'Add Quote',
             'Vendor quote PDF uploaded and parsed into line items',              10),
            ('add_quote_item',       'Add Quote Item',
             'Single line item parsed from a vendor quote (per-item cost)',        1),
            ('add_config',           'Add Base Config',
             'New desired base configuration created for a project',               5),
            ('add_config_component', 'Add Config Component',
             'Component specification added to a base configuration',              1),
            ('archive_quote',        'Archive Quote',
             'Quote hidden from the active view',                                  2),
            ('delete_quote',         'Delete Quote',
             'Quote and all line items permanently deleted',                       2),
            ('compare_quotes',       'Compare Quotes',
             'Side-by-side cost and spec comparison of two vendor quotes',         5),
            ('config_match',         'Config Match',
             'Vendor quote scored against a desired base configuration',          10),
            ('add_tenant',           'Add Tenant',
             'New tenant organisation created',                                    3),
            ('add_project',          'Add Project',
             'New project created under a tenant',                                 3),
            ('archive_tenant',       'Archive Tenant',
             'Tenant and all its projects archived',                               2),
            ('delete_config',        'Delete Config',
             'Base configuration and its component links deleted',                 2),
        ]
        for code, label, desc, cost in _TRANSACTION_TYPES:
            db.execute(
                "INSERT OR IGNORE INTO transaction_types (code,label,description,token_cost) VALUES (?,?,?,?)",
                (code, label, desc, cost)
            )
        db.commit()

        # ── M13b: transactions ledger table ───────────────────────────────────
        if 'transactions' not in tables():
            db.execute("""CREATE TABLE transactions (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                type_id        INTEGER NOT NULL REFERENCES transaction_types(id),
                user_name      TEXT    NOT NULL DEFAULT 'admin',
                quote_id       INTEGER REFERENCES quotes(id)       ON DELETE SET NULL,
                config_id      INTEGER REFERENCES base_configs(id) ON DELETE SET NULL,
                metadata_json  TEXT,
                tokens_charged INTEGER NOT NULL DEFAULT 0,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_transactions_type    ON transactions(type_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_transactions_user    ON transactions(user_name)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_transactions_quote   ON transactions(quote_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_transactions_config  ON transactions(config_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_transactions_created ON transactions(created_at)")
            db.commit()
            logger.info("M13b: transactions ledger table created")

        # ── M14a: servers table ───────────────────────────────────────────────
        if 'servers' not in tables():
            db.execute("""CREATE TABLE servers (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                manufacturer_id INTEGER REFERENCES manufacturers(id),
                model_name      TEXT NOT NULL UNIQUE,
                model_number    TEXT,
                form_factor     TEXT,
                generation      TEXT,
                pdf_path        TEXT,
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_servers_model        ON servers(model_name)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_servers_manufacturer ON servers(manufacturer_id)")
            db.commit()
            logger.info("M14a: servers table created")

        # ── M14b: server_quickspec_components junction ────────────────────────
        if 'server_quickspec_components' not in tables():
            db.execute("""CREATE TABLE server_quickspec_components (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                server_id      INTEGER NOT NULL REFERENCES servers(id)           ON DELETE CASCADE,
                catalog_id     INTEGER NOT NULL REFERENCES component_catalog(id) ON DELETE CASCADE,
                component_role TEXT,
                is_standard    BOOLEAN DEFAULT 1,
                is_optional    BOOLEAN DEFAULT 0,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(server_id, catalog_id))""")
            db.execute("CREATE INDEX IF NOT EXISTS idx_sqc_server  ON server_quickspec_components(server_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_sqc_catalog ON server_quickspec_components(catalog_id)")
            db.commit()
            logger.info("M14b: server_quickspec_components junction created")

        # ── M14c: parse_quickspec transaction type ────────────────────────────
        db.execute("""INSERT OR IGNORE INTO transaction_types (code, label, description, token_cost)
            VALUES ('parse_quickspec', 'Parse QuickSpec',
                    'HPE QuickSpec PDF parsed into server catalog with component options', 5)""")

        # ── M14d: edit_config transaction type ───────────────────────────────
        db.execute("""INSERT OR IGNORE INTO transaction_types (code, label, description, token_cost)
            VALUES ('edit_config', 'Edit Base Config',
                    'Base config name or components updated', 2)""")
        db.commit()

        # ── M15: merge learned_components → defined_components, drop table ────
        if 'learned_components' in tables():
            db.execute("""
                INSERT INTO defined_components (component_type, manufacturer_id, part_number, model, specs)
                SELECT DISTINCT lc.component_type, lc.manufacturer_id, lc.part_number, lc.model, lc.specs_json
                FROM learned_components lc
                WHERE lc.part_number IS NOT NULL AND lc.part_number != ''
                  AND NOT EXISTS (
                      SELECT 1 FROM defined_components dc
                      WHERE dc.part_number = lc.part_number
                        AND dc.component_type = lc.component_type
                  )
            """)
            db.execute("DROP TABLE learned_components")
            db.commit()
            logger.info("M15: learned_components merged into defined_components and dropped")

        # ── M16: merge defined_components → component_catalog, update FK ─────
        # Guard: only run if base_config_components still references defined_components
        # (old schema). If bcc already references component_catalog, M16 has already
        # completed and must not run again — doing so would wipe all component data.
        _bcc_ddl = db.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='base_config_components'"
        ).fetchone()
        _bcc_needs_migration = bool(_bcc_ddl and 'defined_components' in (_bcc_ddl['sql'] or ''))
        if 'defined_components' in tables() and _bcc_needs_migration:
            dc_rows = db.execute("""
                SELECT dc.id, dc.component_type, m.name AS manufacturer,
                       dc.part_number, dc.model, dc.specs
                FROM defined_components dc
                LEFT JOIN manufacturers m ON dc.manufacturer_id = m.id
            """).fetchall()

            id_map = {}
            for row in dc_rows:
                ctype = row['component_type']
                mfr   = row['manufacturer']
                pn    = row['part_number']
                # model is NOT NULL in component_catalog — fall back through part_number → specs
                model = row['model'] or pn or row['specs'] or 'unknown'
                specs = row['specs']

                db.execute("""
                    INSERT OR IGNORE INTO component_catalog
                        (component_type, manufacturer, model, part_number, description, data_source)
                    VALUES (?, ?, ?, ?, ?, 'config')
                """, (ctype, mfr, model, pn, specs))

                if mfr is None:
                    cat = db.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer IS NULL AND model=?",
                        (ctype, model)
                    ).fetchone()
                else:
                    cat = db.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer=? AND model=?",
                        (ctype, mfr, model)
                    ).fetchone()

                if cat:
                    id_map[row['id']] = cat['id']

            # Recreate base_config_components with FK pointing to component_catalog
            db.execute("""
                CREATE TABLE base_config_components_new (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    config_id    INTEGER NOT NULL REFERENCES base_configs(id) ON DELETE CASCADE,
                    component_id INTEGER NOT NULL REFERENCES component_catalog(id),
                    quantity     INTEGER NOT NULL DEFAULT 1
                )
            """)
            for bcc in db.execute(
                "SELECT config_id, component_id, quantity FROM base_config_components"
            ).fetchall():
                new_id = id_map.get(bcc['component_id'])
                if new_id:
                    db.execute(
                        "INSERT INTO base_config_components_new "
                        "(config_id, component_id, quantity) VALUES (?, ?, ?)",
                        (bcc['config_id'], new_id, bcc['quantity'])
                    )
            db.execute("DROP TABLE base_config_components")
            db.execute("ALTER TABLE base_config_components_new RENAME TO base_config_components")
            db.execute("CREATE INDEX IF NOT EXISTS idx_bcc_config    ON base_config_components(config_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_bcc_component ON base_config_components(component_id)")
            db.execute("DROP TABLE defined_components")
            db.commit()
            logger.info("M16: defined_components merged into component_catalog, base_config_components FK updated")

        # ── M17: quotes.config_id ─────────────────────────────────────────────
        if 'config_id' not in cols('quotes'):
            db.execute("ALTER TABLE quotes ADD COLUMN config_id INTEGER REFERENCES base_configs(id) ON DELETE SET NULL")
            db.commit()
            logger.info("M17: quotes.config_id column added")

        # ── M18: normalize quote_date / expiry_date to YYYY-MM-DD ─────────────
        rows = db.execute("SELECT id, quote_date, expiry_date FROM quotes").fetchall()
        updated = 0
        for row in rows:
            new_qd = normalize_date(row['quote_date'])
            new_ed = normalize_date(row['expiry_date'])
            if new_qd != row['quote_date'] or new_ed != row['expiry_date']:
                db.execute(
                    "UPDATE quotes SET quote_date = ?, expiry_date = ? WHERE id = ?",
                    (new_qd, new_ed, row['id'])
                )
                updated += 1
        if updated:
            db.commit()
            logger.info(f"M18: normalized {updated} quote date(s) to YYYY-MM-DD")

        # ── M19: repath quote file path to current DATA_DIR if file moved ───────
        # Column was pdf_path before M25 renamed it to file_path; support both.
        _fcol = 'file_path' if 'file_path' in cols('quotes') else 'pdf_path'
        upload_folder = os.path.join(DATA_DIR, 'uploads')
        pdf_rows = db.execute(f"SELECT id, {_fcol} AS fpath FROM quotes WHERE {_fcol} IS NOT NULL").fetchall()
        repathed = 0
        for row in pdf_rows:
            p = row['fpath']
            if not os.path.exists(p):
                candidate = os.path.join(upload_folder, os.path.basename(p))
                if os.path.exists(candidate):
                    db.execute(f"UPDATE quotes SET {_fcol} = ? WHERE id = ?", (candidate, row['id']))
                    repathed += 1
        if repathed:
            db.commit()
            logger.info(f"M19: repathed {repathed} quote file path(s) to current DATA_DIR")

        # ── M20: repath servers.pdf_path to current DATA_DIR/quickspecs ──────
        quickspec_folder = os.path.join(DATA_DIR, 'quickspecs')
        srv_rows = db.execute("SELECT id, pdf_path FROM servers WHERE pdf_path IS NOT NULL").fetchall()
        srv_repathed = 0
        for row in srv_rows:
            p = row['pdf_path']
            if not os.path.exists(p):
                candidate = os.path.join(quickspec_folder, os.path.basename(p))
                if os.path.exists(candidate):
                    db.execute("UPDATE servers SET pdf_path = ? WHERE id = ?", (candidate, row['id']))
                    srv_repathed += 1
        if srv_repathed:
            db.commit()
            logger.info(f"M20: repathed {srv_repathed} QuickSpec PDF path(s) to current DATA_DIR")

        # ── M21: add quote_items column ───────────────────────────────────────
        if 'quote_items' not in cols('quotes'):
            db.execute("ALTER TABLE quotes ADD COLUMN quote_items INTEGER")
            db.commit()
            logger.info("M21: quotes.quote_items column added")

        # ── M22: default quote_items to 1 where null ──────────────────────────
        updated = db.execute(
            "UPDATE quotes SET quote_items = 1 WHERE quote_items IS NULL"
        ).rowcount
        if updated:
            db.commit()
            logger.info(f"M22: set quote_items = 1 on {updated} quote(s)")

        # ── M23: defined_components (admin-curated approved component list) ───
        if 'defined_components' not in tables():
            db.execute("""
                CREATE TABLE defined_components (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    catalog_id   INTEGER NOT NULL REFERENCES component_catalog(id) ON DELETE CASCADE,
                    label        TEXT,
                    notes        TEXT,
                    is_preferred BOOLEAN NOT NULL DEFAULT 0,
                    created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            db.execute("CREATE INDEX IF NOT EXISTS idx_defined_comp_catalog   ON defined_components(catalog_id)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_defined_comp_preferred ON defined_components(is_preferred)")
            db.execute("""
                CREATE TRIGGER trg_defined_components_updated_at
                AFTER UPDATE ON defined_components
                FOR EACH ROW
                BEGIN
                    UPDATE defined_components SET updated_at = CURRENT_TIMESTAMP WHERE id = OLD.id;
                END
            """)
            db.commit()
            logger.info("M23: defined_components table created")

        # ── M24: users table + seed default accounts ──────────────────────────
        if 'users' not in tables():
            db.execute("""
                CREATE TABLE users (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_name       TEXT NOT NULL UNIQUE,
                    user_email      TEXT,
                    user_gecos      TEXT,
                    password_hash   TEXT NOT NULL,
                    user_last_login TIMESTAMP,
                    user_status     TEXT NOT NULL DEFAULT 'enabled'
                                        CHECK(user_status IN ('enabled', 'disabled')),
                    created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            _pw = generate_password_hash('1q2w3e4r')
            db.execute(
                "INSERT INTO users (user_name, user_email, user_gecos, password_hash) VALUES (?, ?, ?, ?)",
                ('admin@localhost', 'admin@localhost', 'Administrator', _pw)
            )
            db.execute(
                "INSERT INTO users (user_name, user_email, user_gecos, password_hash) VALUES (?, ?, ?, ?)",
                ('toor@localhost', 'toor@localhost', 'Root User', _pw)
            )
            db.execute(
                "INSERT INTO users (user_name, user_email, user_gecos, password_hash) VALUES (?, ?, ?, ?)",
                ('test@test.com', 'test@test.com', 'Test User', _pw)
            )
            db.commit()
            logger.info("M24: users table created, admin and toor accounts seeded")

        # ── M25.5: migrate user_name to email-style; add test user ───────────────
        if 'users' in tables():
            _pw = generate_password_hash('1q2w3e4r')
            # Rename legacy short usernames to their email equivalents
            for old, new_name in [('admin', 'admin@localhost'), ('toor', 'toor@localhost')]:
                row = db.execute("SELECT id FROM users WHERE user_name = ?", (old,)).fetchone()
                if row:
                    db.execute("UPDATE users SET user_name = ? WHERE id = ?", (new_name, row['id']))
            # Add test user if not already present
            if not db.execute("SELECT 1 FROM users WHERE user_name = 'test@test.com'").fetchone():
                db.execute(
                    "INSERT INTO users (user_name, user_email, user_gecos, password_hash) VALUES (?, ?, ?, ?)",
                    ('test@test.com', 'test@test.com', 'Test User', _pw)
                )
            db.commit()
            logger.info("M25.5: user_name migrated to email format, test user added")

        # ── M25: rename quotes.pdf_path → quotes.file_path ────────────────────
        # pdf_path was misleading once Excel uploads were added.
        if 'pdf_path' in cols('quotes') and 'file_path' not in cols('quotes'):
            db.execute("ALTER TABLE quotes RENAME COLUMN pdf_path TO file_path")
            db.commit()
            logger.info("M25: quotes.pdf_path renamed to file_path")

        # ── M26: add projects.delivery_deadline and projects.budget ───────────
        if 'delivery_deadline' not in cols('projects'):
            db.execute("ALTER TABLE projects ADD COLUMN delivery_deadline DATE")
            db.commit()
            logger.info("M26: projects.delivery_deadline column added")
        if 'budget' not in cols('projects'):
            db.execute("ALTER TABLE projects ADD COLUMN budget REAL")
            db.commit()
            logger.info("M26: projects.budget column added")

    finally:
        db.execute("PRAGMA foreign_keys = ON")
        db.close()


def save_quote_to_db(quote_data, line_items, file_path, tenant_id=None, project_id=None, tenant_name='', project_name='', ica='', po_comments='', config_id=None, quote_items=None):
    """Save parsed quote data to database."""
    logger.debug(f"Saving quote: {quote_data.get('quote_id')} with {len(line_items)} items")

    db = get_db()
    cursor = db.cursor()

    try:
        # Resolve quote_id — suffix with -2, -3 … if already taken
        base_quote_id = quote_data.get('quote_id') or 'UNKNOWN'
        quote_id = base_quote_id
        counter = 2
        while cursor.execute('SELECT 1 FROM quotes WHERE quote_id = ?', (quote_id,)).fetchone():
            quote_id = f"{base_quote_id}-{counter}"
            counter += 1

        # Insert quote (with parameterized queries for SQL injection protection)
        cursor.execute('''
            INSERT INTO quotes (quote_id, vendor, customer_name, quote_date, expiry_date,
                               total_amount, currency, description, file_path,
                               tenant_id, project_id, tenant_name, project_name, ica, po_comments, config_id, quote_items)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            quote_id,
            quote_data.get('vendor'),
            quote_data.get('customer_name'),
            normalize_date(quote_data.get('quote_date')),
            normalize_date(quote_data.get('expiry_date')),
            quote_data.get('total_amount'),
            quote_data.get('currency', 'CAD'),
            quote_data.get('description') or os.path.basename(file_path),
            file_path,
            tenant_id,
            project_id,
            tenant_name,
            project_name,
            ica,
            po_comments[:255] if po_comments else None,
            config_id,
            quote_items
        ))

        quote_db_id = cursor.lastrowid

        # Insert line items (simplified without catalog to avoid locks)
        parser = QuoteParser(file_path)
        for item in line_items:
            # Extract component details
            component_details = parser.extract_component_details(item)

            # Insert line item
            cursor.execute('''
                INSERT INTO line_items (quote_id, line_no, quantity, product_number,
                                       description, category, delivery_time)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                quote_db_id,
                item.get('line_no'),
                item.get('quantity'),
                item.get('product_number'),
                item.get('description'),
                item.get('category'),
                item.get('delivery_time')
            ))

        db.commit()
        return quote_db_id

    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()



def get_quote_by_id(quote_db_id):
    """Retrieve quote and all related data with normalized specs."""
    db = get_db()

    # Get quote
    quote = db.execute('SELECT * FROM quotes WHERE id = ?', (quote_db_id,)).fetchone()
    if not quote:
        return None

    # Get line items with catalog info
    line_items = db.execute('''
        SELECT
            li.*,
            cc.id as catalog_id,
            cc.model as catalog_model,
            cc.manufacturer as catalog_manufacturer,
            cc.component_type as catalog_component_type
        FROM line_items li
        LEFT JOIN component_catalog cc ON cc.id = li.catalog_component_id
        WHERE li.quote_id = ?
        ORDER BY li.line_no
    ''', (quote_db_id,)).fetchall()

    db.close()

    # Group by category
    categories = {}
    for item in line_items:
        category = item['category']
        if category not in categories:
            categories[category] = []

        item_dict = dict(item)

        # Try to get normalized specs from catalog if available
        if item_dict.get('catalog_id'):
            normalized_specs = get_normalized_specs(
                item_dict['catalog_id'],
                item_dict.get('catalog_component_type') or category
            )
            item_dict['specs'] = normalized_specs or {}
        else:
            item_dict['specs'] = {}

        # For Memory items without catalog specs, extract capacity_gb from description
        if category == 'Memory' and not item_dict['specs'].get('capacity_gb'):
            desc = item_dict.get('description') or ''
            m = re.search(r'\b(\d+)\s*(TB|GB)\b', desc, re.IGNORECASE)
            if m:
                val, unit = int(m.group(1)), m.group(2).upper()
                item_dict['specs']['capacity_gb'] = val * 1024 if unit == 'TB' else val

        item_dict['urls'] = []

        categories[category].append(item_dict)

    return {
        'quote': dict(quote),
        'categories': categories
    }


def get_normalized_specs(catalog_id, component_type):
    """Get detailed specs from normalized tables."""
    db = get_db()
    specs = None

    try:
        if component_type == 'CPU':
            result = db.execute('''
                SELECT * FROM cpu_specs WHERE catalog_id = ?
            ''', (catalog_id,)).fetchone()

            if result:
                specs = {
                    'cores': result['cores'],
                    'threads': result['threads'],
                    'base_clock_ghz': result['base_clock_ghz'],
                    'max_turbo_clock_ghz': result['max_turbo_clock_ghz'],
                    'l1_cache_kb': result['l1_cache_kb'],
                    'l2_cache_kb': result['l2_cache_kb'],
                    'l3_cache_kb': result['l3_cache_kb'],
                    'cache_mb': result['l3_cache_kb'] / 1024 if result['l3_cache_kb'] else None,
                    'tdp_watts': result['tdp_watts'],
                    'max_memory_gb': result['max_memory_gb'],
                    'memory_channels': result['memory_channels'],
                    'socket': result['socket'],
                    'pcie_lanes': result['pcie_lanes']
                }

        elif component_type == 'Memory':
            result = db.execute('''
                SELECT * FROM memory_specs WHERE catalog_id = ?
            ''', (catalog_id,)).fetchone()

            if result:
                specs = {
                    'capacity_gb': result['capacity_gb'],
                    'speed_mhz': result['speed_mhz'],
                    'ddr_generation': result['ddr_generation'],
                    'module_type': result['module_type'],
                    'ecc_support': bool(result['ecc_support']),
                    'registered': bool(result['registered'])
                }

        elif component_type == 'Disk':
            result = db.execute('''
                SELECT * FROM disk_specs WHERE catalog_id = ?
            ''', (catalog_id,)).fetchone()

            if result:
                specs = {
                    'capacity_gb': result['capacity_gb'],
                    'capacity_tb': result['capacity_tb'],
                    'disk_type': result['disk_type'],
                    'interface': result['interface'],
                    'read_speed_mbps': result['read_speed_mbps'],
                    'write_speed_mbps': result['write_speed_mbps'],
                    'iops_read': result['iops_read'],
                    'iops_write': result['iops_write']
                }

        elif component_type == 'Network Card':
            result = db.execute('''
                SELECT * FROM network_card_specs WHERE catalog_id = ?
            ''', (catalog_id,)).fetchone()

            if result:
                specs = {
                    'port_count': result['port_count'],
                    'speed_gbps': result['speed_gbps'],
                    'port_type': result['port_type'],
                    'rdma_support': bool(result['rdma_support'])
                }

    finally:
        db.close()

    return specs


def log_transaction(type_code, user_name=None, quote_id=None, config_id=None,
                    metadata=None, tokens_override=None):
    """
    Append one row to the transactions ledger.
    Never raises — logging failures must not break the calling API.

    tokens_override: when provided, overrides transaction_types.token_cost.
    Useful for compound actions (e.g. add_quote + N × add_quote_item).
    """
    if user_name is None:
        user_name = _current_user()
    try:
        db = get_db()
        row = db.execute(
            "SELECT id, token_cost FROM transaction_types WHERE code = ?", (type_code,)
        ).fetchone()
        if not row:
            logger.warning(f"log_transaction: unknown type_code '{type_code}'")
            db.close()
            return
        tokens = tokens_override if tokens_override is not None else row['token_cost']
        db.execute(
            """INSERT INTO transactions
               (type_id, user_name, quote_id, config_id, metadata_json, tokens_charged)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (row['id'], user_name, quote_id, config_id,
             json.dumps(metadata) if metadata else None, tokens)
        )
        db.commit()
        db.close()
    except Exception as e:
        logger.warning(f"log_transaction failed for '{type_code}': {e}")


def get_all_quotes():
    """Get all quotes for listing."""
    db = get_db()
    quotes = db.execute('''
        SELECT id, quote_id, vendor, customer_name, quote_date, expiry_date,
               total_amount, currency, description, tenant_name, project_name,
               ica, status, po_comments, uploaded_at, config_id, quote_items
        FROM quotes
        WHERE status != 'archived'
        ORDER BY uploaded_at DESC
    ''').fetchall()
    db.close()
    return [dict(q) for q in quotes]


# =============================================================================
# AUTH ROUTES
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('user_name'):
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE user_name = ?", (username,)
        ).fetchone()
        db.close()
        if user and user['user_status'] == 'enabled' and check_password_hash(user['password_hash'], password):
            session.permanent = True
            session['user_name'] = user['user_name']
            db2 = get_db()
            db2.execute(
                "UPDATE users SET user_last_login = CURRENT_TIMESTAMP WHERE id = ?", (user['id'],)
            )
            db2.commit()
            db2.close()
            return redirect(url_for('index'))
        elif user and user['user_status'] == 'disabled':
            error = 'Account is disabled.'
        else:
            error = 'Invalid username or password.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.before_request
def require_login():
    """Block all routes except /login and static assets when not authenticated."""
    public = {'/login', '/logout'}
    if request.path in public or request.path.startswith('/static/'):
        return None
    if not session.get('user_name'):
        return redirect(url_for('login'))


@app.route('/')
def index():
    """Main page with tenant/project hierarchy navigation."""
    import re as _re
    db = get_db()
    _nic_re = _re.compile(
        r'ethernet|infiniband|adapter|omni-path|\bnic\b|gbe\b|10g|25g|40g|100g|'
        r'sfp|qsfp|osfp|ocp\s*\d|\bfcoe\b|\broce\b',
        _re.IGNORECASE
    )
    rows = db.execute("""
        SELECT DISTINCT cc.part_number, cc.description, cc.manufacturer
        FROM component_catalog cc
        JOIN server_quickspec_components sqc ON sqc.catalog_id = cc.id
        WHERE cc.component_type = 'Network Card'
        ORDER BY cc.manufacturer, cc.part_number
    """).fetchall()
    db.close()
    raw = [dict(r) for r in rows if r['description'] and _nic_re.search(r['description'])]
    seen = {}
    for r in raw:
        pn = r['part_number']
        if pn not in seen or seen[pn]['manufacturer'] in (None, 'Unknown'):
            seen[pn] = r
    nic_catalog = sorted(seen.values(), key=lambda r: (r['manufacturer'] or '', r['part_number']))
    return render_template('index_new.html', nic_catalog=nic_catalog)


@app.route('/quote/<int:quote_id>')
def view_quote(quote_id):
    """View single quote card."""
    quote_data = get_quote_by_id(quote_id)
    if not quote_data:
        return "Quote not found", 404

    # Check if split view is requested (default to split)
    use_split_view = request.args.get('view', 'split') == 'split'

    if use_split_view:
        return render_template('quote_card_split.html', data=quote_data)
    else:
        return render_template('quote_card.html', data=quote_data)


@app.route('/upload', methods=['POST'])
def upload_quote():
    """Upload and parse a PDF or Dell Excel (.xlsx) quote."""
    logger.info(f"Upload request received from {request.remote_addr}")

    if 'file' not in request.files:
        logger.warning("Upload rejected: No file in request")
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        logger.warning("Upload rejected: Empty filename")
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        logger.warning(f"Upload rejected: Invalid file type - {file.filename}")
        return jsonify({'error': 'Only PDF and Excel (.xlsx) files allowed'}), 400

    ext = file.filename.rsplit('.', 1)[1].lower()

    logger.info(f"Processing upload: {file.filename}")

    try:
        # Get metadata first so we can build the filename
        tenant_id = request.form.get('tenant_id', '').strip()
        project_id = request.form.get('project_id', '').strip()
        ica = request.form.get('ica', '').strip()
        po_comments = request.form.get('po_comments', '').strip()
        config_id_raw = request.form.get('config_id', '').strip()
        quote_items_raw = request.form.get('quote_items', '').strip()

        tenant_id = int(tenant_id) if tenant_id else None
        project_id = int(project_id) if project_id else None
        config_id = int(config_id_raw) if config_id_raw else None
        quote_items = int(quote_items_raw) if quote_items_raw and quote_items_raw.isdigit() else None

        # Look up tenant/project names for the filename
        db = get_db()
        tenant_name = ''
        project_name = ''
        if tenant_id:
            row = db.execute('SELECT name FROM tenants WHERE id = ?', (tenant_id,)).fetchone()
            if row:
                tenant_name = row['name']
        if project_id:
            row = db.execute('SELECT name FROM projects WHERE id = ?', (project_id,)).fetchone()
            if row:
                project_name = row['name']
        db.close()

        # Build filename: {tenant}-{project}-{ica}-{original}
        parts = [tenant_name, project_name, ica]
        prefix = '-'.join(secure_filename(p) for p in parts if p)
        original_name = secure_filename(file.filename)
        filename = f"{prefix}-{original_name}" if prefix else original_name
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        if ext == 'pdf':
            # Validate PDF security
            logger.info(f"Validating PDF: {filename}")
            is_valid, error_msg = validate_pdf(filepath)
            if not is_valid:
                logger.warning(f"PDF validation failed for {filename}: {error_msg}")
                os.remove(filepath)
                return jsonify({'error': f'PDF validation failed: {error_msg}'}), 400
            logger.info(f"Parsing PDF: {filename}")
            parser = QuoteParser(filepath)
        else:
            # Excel quote (Dell format)
            logger.info(f"Parsing Excel quote: {filename}")
            parser = DellExcelParser(filepath)

        result = parser.parse()
        logger.info(f"Parsed {len(result.get('line_items', []))} line items from {filename}")

        # Save to database
        logger.info(f"Saving quote to database: {result['quote'].get('quote_id', 'Unknown')}")
        quote_db_id = save_quote_to_db(
            result['quote'],
            result['line_items'],
            filepath,
            tenant_id=tenant_id,
            project_id=project_id,
            tenant_name=tenant_name,
            project_name=project_name,
            ica=ica,
            po_comments=po_comments,
            config_id=config_id,
            quote_items=quote_items
        )

        logger.info(f"Quote saved successfully with ID: {quote_db_id}")

        item_count = len(result.get('line_items', []))
        # Base cost 10 (add_quote) + 1 per line item (add_quote_item)
        log_transaction(
            'add_quote',
            quote_id=quote_db_id,
            metadata={
                'filename': filename,
                'vendor': result['quote'].get('vendor'),
                'item_count': item_count,
                'tenant_id': tenant_id,
                'project_id': project_id,
            },
            tokens_override=10 + item_count
        )

        return jsonify({
            'success': True,
            'quote_id': quote_db_id,
            'redirect': url_for('view_quote', quote_id=quote_db_id)
        })

    except Exception as e:
        logger.error(f"Error processing upload {file.filename}: {e}", exc_info=True)
        # Clean up uploaded file on error
        if 'filepath' in locals() and os.path.exists(filepath):
            try:
                os.remove(filepath)
                logger.info(f"Cleaned up failed upload: {filepath}")
            except Exception as cleanup_error:
                logger.error(f"Failed to clean up file {filepath}: {cleanup_error}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/quotes')
def api_quotes():
    """API endpoint for quotes list."""
    quotes = get_all_quotes()
    return jsonify(quotes)


@app.route('/api/quote/<int:quote_id>')
def api_quote(quote_id):
    """API endpoint for single quote."""
    quote_data = get_quote_by_id(quote_id)
    if not quote_data:
        return jsonify({'error': 'Quote not found'}), 404
    return jsonify(quote_data)


def _excel_to_html(path):
    """Render an xlsx workbook as a self-contained styled HTML page.

    Handles:
    - Merged cells (colspan/rowspan via HTML attributes)
    - Uncached formula cells (load both cached and raw workbook, prefer cached)
    - Wide sheets (horizontal scroll wrapper)
    """
    import openpyxl

    # Load twice: data_only for cached formula results, and raw for formula text.
    # Cells whose cached value is None but have a formula get a '(formula)' hint
    # so they don't silently appear blank.
    wb_data = openpyxl.load_workbook(path, data_only=True)
    wb_raw  = openpyxl.load_workbook(path, data_only=False)
    ws_data = wb_data.active
    ws_raw  = wb_raw.active

    # Build a set of merged ranges and a skip-set for cells that are
    # covered (non-top-left) by a merge — those must not emit a <td>.
    # openpyxl exposes ws.merged_cells.ranges
    merge_attrs = {}   # (row, col) -> {'rowspan': r, 'colspan': c}
    skip_cells  = set()

    for merge in ws_data.merged_cells.ranges:
        min_r, min_c = merge.min_row, merge.min_col
        max_r, max_c = merge.max_row, merge.max_col
        rowspan = max_r - min_r + 1
        colspan = max_c - min_c + 1
        merge_attrs[(min_r, min_c)] = {'rowspan': rowspan, 'colspan': colspan}
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    skip_cells.add((r, c))

    # Find actual used bounds (trim trailing fully-empty rows)
    all_rows = list(ws_data.iter_rows())
    while all_rows and all(cell.value is None for cell in all_rows[-1]):
        all_rows.pop()
    if not all_rows:
        return '<html><body><p>Empty spreadsheet.</p></body></html>'

    max_col = max(len(row) for row in all_rows)

    def _fmt(cell_data, cell_raw):
        v = cell_data.value
        if v is None:
            # Check if the raw cell has a formula — if so value was never cached
            rv = cell_raw.value if cell_raw else None
            if rv and str(rv).startswith('='):
                return '<span style="color:#94a3b8;font-style:italic">(formula)</span>'
            return ''
        if isinstance(v, float):
            return f'{v:,.2f}' if v != int(v) else f'{int(v):,}'
        return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    body_rows = ''
    for row_data in all_rows:
        row_idx = row_data[0].row
        cells_html = ''
        for cell_data in row_data[:max_col]:
            col_idx = cell_data.column
            coord    = (row_idx, col_idx)

            if coord in skip_cells:
                continue  # covered by a merge — omit

            attrs = merge_attrs.get(coord, {})
            span_attrs = ''
            if attrs.get('rowspan', 1) > 1:
                span_attrs += f' rowspan="{attrs["rowspan"]}"'
            if attrs.get('colspan', 1) > 1:
                span_attrs += f' colspan="{attrs["colspan"]}"'

            try:
                cell_raw = ws_raw.cell(row=row_idx, column=col_idx)
            except Exception:
                cell_raw = None

            content = _fmt(cell_data, cell_raw)
            cells_html += f'<td{span_attrs}>{content}</td>'

        body_rows += f'<tr>{cells_html}</tr>\n'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    font-size: 11px;
    color: #1a1a2e;
    background: #f8fafc;
    padding: 16px;
  }}
  .scroll-wrapper {{
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    border-radius: 8px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
  }}
  table {{
    border-collapse: collapse;
    min-width: 100%;
    background: #fff;
  }}
  th, td {{
    padding: 5px 8px;
    border: 1px solid #e2e8f0;
    white-space: pre-wrap;
    word-break: break-word;
    vertical-align: top;
  }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  tr:hover td {{ background: #ede9fe; }}
</style>
</head>
<body>
<div class="scroll-wrapper">
<table>
<tbody>
{body_rows}
</tbody>
</table>
</div>
</body>
</html>"""


@app.route('/pdf/<int:quote_id>')
def serve_pdf(quote_id):
    """Serve the source file for a quote — PDF inline or Excel rendered as HTML."""
    from flask import send_file, make_response
    db = get_db()
    quote = db.execute('SELECT file_path FROM quotes WHERE id = ?', (quote_id,)).fetchone()
    db.close()
    if not quote or not quote['file_path']:
        return "Document not found", 404
    file_path = quote['file_path']
    if not os.path.isabs(file_path):
        file_path = os.path.join(_app_dir, file_path)
    if not os.path.exists(file_path):
        fallback = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(file_path))
        if os.path.exists(fallback):
            file_path = fallback
        else:
            return "File not found on disk", 404

    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        html = _excel_to_html(file_path)
        resp = make_response(html)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp

    return send_file(file_path, mimetype='application/pdf')


@app.route('/compare')
def compare_quotes():
    """Compare two quotes side by side."""
    ids_param = request.args.get('ids', '')

    if not ids_param:
        return "No quotes selected for comparison", 400

    try:
        quote_ids = [int(id.strip()) for id in ids_param.split(',')]
    except ValueError:
        return "Invalid quote IDs", 400

    if len(quote_ids) != 2:
        return "Please select exactly 2 quotes to compare", 400

    # Fetch both quotes
    quote1_data = get_quote_by_id(quote_ids[0])
    quote2_data = get_quote_by_id(quote_ids[1])

    if not quote1_data or not quote2_data:
        return "One or both quotes not found", 404

    # Get union of all categories from both quotes, preserving order
    category_order = ['CPU', 'Memory', 'Disk', 'Storage Controller',
                      'Network Card', 'Power Supply', 'GPU', 'Additional Hardware']

    all_categories = [cat for cat in category_order
                      if cat in quote1_data['categories'] or cat in quote2_data['categories']]

    log_transaction(
        'compare_quotes',
        metadata={'quote_ids': quote_ids,
                  'quote_refs': [quote1_data['quote']['quote_id'],
                                 quote2_data['quote']['quote_id']]}
    )

    # Check if split view is requested
    use_split_view = request.args.get('view', 'split') == 'split'

    if use_split_view:
        return render_template('compare_split.html',
                              quote1=quote1_data,
                              quote2=quote2_data,
                              categories=all_categories)
    else:
        return render_template('compare.html',
                              quote1=quote1_data,
                              quote2=quote2_data,
                              categories=all_categories)


# =============================================================================
# ADMIN ROUTES
# =============================================================================

@app.route('/admin/users')
def admin_users():
    db = get_db()
    users = db.execute(
        "SELECT id, user_name, user_email, user_gecos, user_last_login, user_status, created_at FROM users ORDER BY user_name"
    ).fetchall()
    db.close()
    return render_template('admin/users_list.html', users=[dict(u) for u in users])


@app.route('/admin/users/<int:user_id>/edit')
def admin_user_edit(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    db.close()
    if not user:
        return "User not found", 404
    return render_template('admin/user_edit.html', user=dict(user))


@app.route('/admin/users/<int:user_id>/update', methods=['POST'])
def admin_user_update(user_id):
    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        db.close()
        return "User not found", 404

    user_email  = request.form.get('user_email', '').strip()
    user_gecos  = request.form.get('user_gecos', '').strip()
    user_status = request.form.get('user_status', 'enabled')
    new_password = request.form.get('new_password', '').strip()

    if user_status not in ('enabled', 'disabled'):
        user_status = 'enabled'

    if new_password:
        pw_hash = generate_password_hash(new_password)
        db.execute(
            "UPDATE users SET user_email=?, user_gecos=?, user_status=?, password_hash=? WHERE id=?",
            (user_email, user_gecos, user_status, pw_hash, user_id)
        )
    else:
        db.execute(
            "UPDATE users SET user_email=?, user_gecos=?, user_status=? WHERE id=?",
            (user_email, user_gecos, user_status, user_id)
        )
    db.commit()
    db.close()
    return redirect(url_for('admin_dashboard') + '?view=users')


@app.route('/admin/users/create', methods=['POST'])
def admin_user_create():
    user_email  = request.form.get('user_email', '').strip()
    user_gecos  = request.form.get('user_gecos', '').strip()
    user_status = request.form.get('user_status', 'enabled')
    password    = request.form.get('new_password', '').strip()

    # Username = email; fall back to user@local if blank
    user_name = user_email or 'user@local'
    if user_status not in ('enabled', 'disabled'):
        user_status = 'enabled'
    if not password:
        return jsonify({'error': 'Password is required'}), 400

    pw_hash = generate_password_hash(password)
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (user_name, user_email, user_gecos, password_hash, user_status) VALUES (?, ?, ?, ?, ?)",
            (user_name, user_email, user_gecos, pw_hash, user_status)
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        db.close()


@app.route('/admin')
def admin_dashboard():
    """Admin dashboard with split pane layout."""
    return render_template('admin/dashboard.html')


@app.route('/admin/tenants')
def admin_tenants():
    """List all active tenants."""
    db = get_db()
    tenants = db.execute('''
        SELECT id, name, contact_name, created_at, updated_at, status
        FROM tenants
        WHERE status != 'archived'
        ORDER BY name
    ''').fetchall()
    db.close()
    return render_template('admin/tenants_list.html', tenants=[dict(t) for t in tenants])


@app.route('/admin/tenants/archived')
def admin_tenants_archived():
    """List all archived tenants."""
    db = get_db()
    tenants = db.execute('''
        SELECT id, name, contact_name, created_at, updated_at, status
        FROM tenants
        WHERE status = 'archived'
        ORDER BY name
    ''').fetchall()
    db.close()
    return render_template('admin/archived_tenants_list.html', tenants=[dict(t) for t in tenants])


@app.route('/admin/tenants/<int:tenant_id>/edit')
def admin_tenant_edit(tenant_id):
    """Show edit form for a tenant."""
    db = get_db()
    tenant = db.execute(
        'SELECT id, name, contact_name, status FROM tenants WHERE id = ?', (tenant_id,)
    ).fetchone()
    db.close()
    if not tenant:
        return "Tenant not found", 404
    return render_template('admin/tenant_edit.html', tenant=dict(tenant))


@app.route('/admin/tenants/<int:tenant_id>/update', methods=['POST'])
def admin_tenant_update(tenant_id):
    """Save tenant edits and redirect to admin."""
    tenant_name = request.form.get('tenant_name', '').strip()
    contact_name = request.form.get('contact_name', '').strip()

    if not tenant_name:
        return "Tenant name is required", 400

    db = get_db()
    try:
        db.execute('''
            UPDATE tenants
            SET name = ?, contact_name = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (tenant_name, contact_name or None, tenant_id))
        db.commit()
        logger.info(f"Updated tenant {tenant_id}: name={tenant_name}, contact={contact_name}")
    except sqlite3.IntegrityError:
        return "A tenant with that name already exists", 400
    finally:
        db.close()

    return redirect(url_for('admin_dashboard'))


@app.route('/api/admin/tenants/<int:tenant_id>/archive', methods=['POST'])
def api_admin_tenant_archive(tenant_id):
    """Archive a tenant and all its projects."""
    db = get_db()
    try:
        db.execute(
            "UPDATE projects SET status = 'archived', updated_at = CURRENT_TIMESTAMP WHERE tenant_id = ?",
            (tenant_id,)
        )
        db.execute(
            "UPDATE tenants SET status = 'archived', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (tenant_id,)
        )
        db.commit()
        logger.info(f"Archived tenant {tenant_id} and all its projects")
        log_transaction('archive_tenant', metadata={'tenant_id': tenant_id})
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        logger.error(f"Error archiving tenant {tenant_id}: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/admin/tenants/new')
def admin_tenant_new():
    """Show form to create new tenant."""
    return render_template('admin/tenant_form.html')


@app.route('/admin/tenants/create', methods=['POST'])
def admin_tenant_create():
    """Create a new tenant."""
    tenant_name = request.form.get('tenant_name', '').strip()

    if not tenant_name:
        return jsonify({'error': 'Tenant name is required'}), 400

    db = get_db()
    try:
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO tenants (name, status)
            VALUES (?, 'active')
        ''', (tenant_name,))
        db.commit()
        tenant_id = cursor.lastrowid
        logger.info(f"Created tenant: {tenant_name} (ID: {tenant_id})")
        log_transaction('add_tenant', metadata={'tenant_id': tenant_id, 'name': tenant_name})
        return jsonify({'success': True, 'tenant_id': tenant_id})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Tenant with this name already exists'}), 400
    except Exception as e:
        logger.error(f"Error creating tenant: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/admin/tenants')
def api_admin_tenants():
    """API endpoint for active tenants list."""
    db = get_db()
    tenants = db.execute('''
        SELECT id, name, created_at, updated_at, status
        FROM tenants
        WHERE status != 'archived'
        ORDER BY name
    ''').fetchall()
    db.close()
    return jsonify([dict(t) for t in tenants])


# =============================================================================
# NAVIGATION API ROUTES
# =============================================================================

@app.route('/api/navigation/hierarchy')
def api_navigation_hierarchy():
    """Get complete tenant -> project -> quote hierarchy."""
    db = get_db()

    # Get all tenants with their projects and quote counts
    tenants = db.execute('''
        SELECT
            t.id,
            t.name,
            t.status,
            COUNT(DISTINCT p.id) as project_count
        FROM tenants t
        LEFT JOIN projects p ON p.tenant_id = t.id
        WHERE t.status = 'active'
        GROUP BY t.id
        ORDER BY t.name
    ''').fetchall()

    hierarchy = []
    for tenant in tenants:
        tenant_dict = dict(tenant)

        # Get projects for this tenant with quote counts
        projects = db.execute('''
            SELECT
                p.id,
                p.name,
                p.description,
                p.delivery_deadline,
                p.budget,
                p.status,
                COUNT(q.id) as quote_count
            FROM projects p
            LEFT JOIN quotes q ON q.project_id = p.id
            WHERE p.tenant_id = ? AND p.status != 'archived'
            GROUP BY p.id
            ORDER BY p.name
        ''', (tenant['id'],)).fetchall()

        project_list = []
        for p in projects:
            project_dict = dict(p)
            vendors = db.execute('''
                SELECT vendor, COUNT(id) as count
                FROM quotes
                WHERE project_id = ? AND status != 'archived'
                  AND vendor IS NOT NULL AND vendor != ''
                GROUP BY vendor
                ORDER BY vendor
            ''', (p['id'],)).fetchall()
            project_dict['vendors'] = [dict(v) for v in vendors]
            project_list.append(project_dict)

        tenant_dict['projects'] = project_list
        hierarchy.append(tenant_dict)

    db.close()
    return jsonify(hierarchy)


@app.route('/api/projects/<int:project_id>/quotes')
def api_project_quotes(project_id):
    """Get all quotes for a specific project."""
    db = get_db()
    quotes = db.execute('''
        SELECT q.id, q.quote_id, q.vendor, q.customer_name, q.quote_date, q.expiry_date,
               q.total_amount, q.currency, q.description, q.tenant_name, q.project_name,
               q.ica, q.status, q.po_comments, q.uploaded_at, q.config_id, q.quote_items,
               bc.config_name
        FROM quotes q
        LEFT JOIN base_configs bc ON bc.id = q.config_id
        WHERE q.project_id = ? AND q.status != 'archived'
        ORDER BY q.uploaded_at DESC
    ''', (project_id,)).fetchall()
    result = []
    for q in quotes:
        row = dict(q)
        row['component_tags'] = _build_component_tags(db, q['id'])
        result.append(row)
    db.close()
    return jsonify(result)


def _build_component_tags(db, quote_id):
    """Return a list of compact component tag dicts for a quote's line items."""
    rows = db.execute('''
        SELECT category, description, quantity
        FROM line_items
        WHERE quote_id = ?
        ORDER BY CASE category
            WHEN 'CPU' THEN 1
            WHEN 'Memory' THEN 2
            WHEN 'Disk' THEN 3
            WHEN 'Network Card' THEN 4
            ELSE 5
        END
    ''', (quote_id,)).fetchall()

    tags = []
    seen_cats = set()
    for row in rows:
        cat = row['category']
        if cat not in ('CPU', 'Memory', 'Disk', 'Network Card'):
            continue
        if cat in seen_cats:
            continue
        seen_cats.add(cat)
        desc = (row['description'] or '').strip()
        qty = row['quantity'] or 1

        # Extract compact label from description
        # Remove vendor prefixes like "HPE ", "Dell ", "Intel ", etc.
        label = desc
        # Strip common vendor/product prefixes
        label = re.sub(r'^(?:HPE|Dell|AMD|Intel|Broadcom|INT|BCM|Seagate|Samsung|Micron|Kingston)\s+', '', label, flags=re.IGNORECASE)
        # For CPU: grab model fragment (e.g. "EPYC 9534", "Xeon Platinum 8592+")
        if cat == 'CPU':
            m = re.search(r'(EPYC\s+\w+|Xeon\s+\w+(?:\s+\w+)?|\w+-\d+\w*)', label, re.IGNORECASE)
            label = m.group(1) if m else label.split()[0] if label else desc
        # For Memory: grab capacity fragment (e.g. "96GB", "16GB RDIMM")
        elif cat == 'Memory':
            m = re.search(r'(\d+\s*[GT]B(?:\s+\w+)?)', label, re.IGNORECASE)
            label = m.group(1).strip() if m else label.split()[0] if label else desc
        # For Disk: grab capacity + type
        elif cat == 'Disk':
            # Skip pure boot/management devices
            if re.search(r'\bBOSS\b', desc, re.IGNORECASE):
                seen_cats.discard(cat)
                continue
            m = re.search(r'(\d+(?:\.\d+)?\s*[GT]B(?:\s+\w+)?)', label, re.IGNORECASE)
            label = m.group(1).strip() if m else label.split()[0] if label else desc
        # For NIC: grab speed fragment
        elif cat == 'Network Card':
            m = re.search(r'(\d+\s*G[bB]e?(?:\s+\w+)?)', label, re.IGNORECASE)
            label = m.group(1).strip() if m else label.split()[0] if label else desc

        tags.append({'cat': cat, 'label': f'{qty}\u00d7 {label}'})

    return tags


@app.route('/api/quotes/unassigned')
def api_unassigned_quotes():
    """Get quotes not assigned to any project."""
    db = get_db()
    quotes = db.execute('''
        SELECT q.id, q.quote_id, q.vendor, q.customer_name, q.quote_date, q.expiry_date,
               q.total_amount, q.currency, q.description, q.tenant_name, q.project_name,
               q.ica, q.status, q.po_comments, q.uploaded_at, q.config_id, q.quote_items,
               bc.config_name
        FROM quotes q
        LEFT JOIN base_configs bc ON bc.id = q.config_id
        WHERE q.project_id IS NULL AND q.status != 'archived'
        ORDER BY q.uploaded_at DESC
    ''').fetchall()
    result = []
    for q in quotes:
        row = dict(q)
        row['component_tags'] = _build_component_tags(db, q['id'])
        result.append(row)
    db.close()
    return jsonify(result)


# =============================================================================
# PROJECT MANAGEMENT ROUTES
# =============================================================================

@app.route('/admin/projects')
def admin_projects():
    """List all projects grouped by tenant."""
    db = get_db()

    # Get all active tenants
    tenants = db.execute("SELECT id, name, status FROM tenants WHERE status != 'archived' ORDER BY name").fetchall()

    # Build tenant-project hierarchy
    tenant_projects = []
    for tenant in tenants:
        tenant_dict = dict(tenant)

        # Get projects for this tenant
        projects = db.execute('''
            SELECT p.id, p.name, p.description, p.comments, p.status, p.created_at,
                   COUNT(q.id) as quote_count
            FROM projects p
            LEFT JOIN quotes q ON q.project_id = p.id
            WHERE p.tenant_id = ? AND p.status != 'archived'
            GROUP BY p.id
            ORDER BY p.name
        ''', (tenant['id'],)).fetchall()

        tenant_dict['projects'] = [dict(p) for p in projects]
        tenant_dict['project_count'] = len(projects)
        tenant_projects.append(tenant_dict)

    db.close()
    return render_template('admin/projects_list.html', tenant_projects=tenant_projects)


@app.route('/admin/quotes')
def admin_quotes():
    """List all quotes in the system."""
    db = get_db()
    quotes = db.execute("""
        SELECT id, quote_id, vendor, quote_date, expiry_date,
               total_amount, currency, quote_items,
               CASE WHEN quote_items > 0 THEN ROUND(CAST(total_amount AS REAL) / quote_items, 2) ELSE NULL END AS unit_cost,
               tenant_name, project_name, ica, uploaded_at, status, po_comments
        FROM quotes
        ORDER BY uploaded_at DESC
    """).fetchall()
    db.close()
    return render_template('admin/quotes_list.html', quotes=[dict(q) for q in quotes])


@app.route('/api/admin/quotes/<int:quote_id>/delete', methods=['POST'])
def api_admin_quote_delete(quote_id):
    """Delete a quote and all associated data."""
    db = get_db()
    try:
        db.execute("PRAGMA foreign_keys = ON")
        row = db.execute("SELECT file_path FROM quotes WHERE id = ?", (quote_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Quote not found'}), 404
        db.execute("DELETE FROM quotes WHERE id = ?", (quote_id,))
        db.commit()
        log_transaction('delete_quote', metadata={'deleted_quote_id': quote_id})
        # Remove uploaded file if it exists
        if row['file_path'] and os.path.exists(row['file_path']):
            try:
                os.remove(row['file_path'])
            except OSError:
                pass
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/quotes/<int:quote_id>/items', methods=['PATCH'])
def api_quote_update_items(quote_id):
    """Update quote_items for a quote."""
    data = request.get_json(silent=True) or {}
    try:
        qty = int(data.get('quote_items', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid value'}), 400
    if qty < 1 or qty > 255:
        return jsonify({'error': 'Quantity must be between 1 and 255'}), 400
    db = get_db()
    try:
        if not db.execute("SELECT id FROM quotes WHERE id = ?", (quote_id,)).fetchone():
            return jsonify({'error': 'Quote not found'}), 404
        db.execute("UPDATE quotes SET quote_items = ? WHERE id = ?", (qty, quote_id))
        db.commit()
        return jsonify({'success': True, 'quote_items': qty})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/quotes/<int:quote_id>/archive', methods=['POST'])
def api_quote_archive(quote_id):
    """Archive a quote (hides it from the frontend)."""
    db = get_db()
    try:
        row = db.execute("SELECT id FROM quotes WHERE id = ?", (quote_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Quote not found'}), 404
        db.execute("UPDATE quotes SET status = 'archived' WHERE id = ?", (quote_id,))
        db.commit()
        log_transaction('archive_quote', quote_id=quote_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# =============================================================================
# MANUFACTURERS API
# =============================================================================

@app.route('/api/manufacturers')
def api_manufacturers():
    db = get_db()
    rows = db.execute("SELECT id, name FROM manufacturers ORDER BY name").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# =============================================================================
# COMPONENTS API (read-only, for dropdown population)
# =============================================================================

@app.route('/api/components/learned')
def api_learned_components():
    ctype = request.args.get('type', '').strip()
    q     = request.args.get('q', '').strip()

    sql    = "SELECT component_type, manufacturer, part_number, model, description FROM component_catalog WHERE 1=1"
    params = []
    if ctype:
        sql += " AND component_type = ?"
        params.append(ctype)
    if q:
        like = f'%{q}%'
        sql += " AND (model LIKE ? OR part_number LIKE ? OR description LIKE ? OR manufacturer LIKE ?)"
        params.extend([like, like, like, like])
    sql += " ORDER BY manufacturer, part_number LIMIT 15"

    db   = get_db()
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# =============================================================================
# BASE CONFIGS API
# =============================================================================

@app.route('/api/projects/<int:project_id>/configs')
def api_project_configs(project_id):
    db = get_db()
    try:
        rows = db.execute("""
            SELECT bc.id, bc.config_name, bc.created_at,
                   cc.id AS comp_id, cc.component_type, cc.part_number,
                   cc.description AS specs, cc.model,
                   cc.manufacturer AS manufacturer_name, bcc.quantity
            FROM base_configs bc
            LEFT JOIN base_config_components bcc ON bcc.config_id = bc.id
            LEFT JOIN component_catalog cc ON bcc.component_id = cc.id
            WHERE bc.project_id = ?
            ORDER BY bc.created_at DESC, cc.component_type, cc.part_number
        """, (project_id,)).fetchall()
        configs, order = {}, []
        for r in rows:
            cid = r['id']
            if cid not in configs:
                configs[cid] = {'id': cid, 'config_name': r['config_name'],
                                'created_at': r['created_at'], 'components': []}
                order.append(cid)
            if r['comp_id'] is not None:
                configs[cid]['components'].append({
                    'id': r['comp_id'], 'component_type': r['component_type'],
                    'part_number': r['part_number'], 'specs': r['specs'],
                    'model': r['model'], 'manufacturer_name': r['manufacturer_name'],
                    'quantity': r['quantity'],
                })
        return jsonify([configs[cid] for cid in order])
    except Exception as e:
        logger.error(f"api_project_configs error: {e}")
        return jsonify({'error': 'Failed to load configs'}), 500
    finally:
        db.close()


@app.route('/api/projects/<int:project_id>/configs', methods=['POST'])
def api_create_config(project_id):
    data = request.get_json()
    if not data or not data.get('config_name'):
        return jsonify({'error': 'config_name required'}), 400
    db = get_db()
    try:
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO base_configs (config_name, project_id) VALUES (?, ?)",
            (data['config_name'].strip(), project_id)
        )
        config_id = cursor.lastrowid
        for comp in data.get('components', []):
            ctype = comp.get('component_type')
            pn    = comp.get('part_number') or None
            model = comp.get('model') or pn or comp.get('specs') or 'unknown'
            specs = comp.get('specs') or None

            # Accept manufacturer as text directly, or resolve from id
            mfr = comp.get('manufacturer') or None
            if not mfr:
                mfr_id = comp.get('manufacturer_id') or None
                if mfr_id:
                    mfr_row = cursor.execute(
                        "SELECT name FROM manufacturers WHERE id=?", (mfr_id,)
                    ).fetchone()
                    if mfr_row:
                        mfr = mfr_row['name']

            # Upsert into component_catalog, then link via junction
            cursor.execute("""
                INSERT OR IGNORE INTO component_catalog
                    (component_type, manufacturer, model, part_number, description, data_source)
                VALUES (?, ?, ?, ?, ?, 'config')
            """, (ctype, mfr, model, pn, specs))

            if cursor.rowcount == 1:
                # Fresh insert — use the new row directly (avoids returning a stale
                # row with a different part_number when manufacturer IS NULL)
                cat_id = cursor.lastrowid
            else:
                # Row already existed (UNIQUE conflict on non-NULL manufacturer) — find it
                if mfr is None:
                    row = cursor.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer IS NULL AND model=?",
                        (ctype, model)
                    ).fetchone()
                else:
                    row = cursor.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer=? AND model=?",
                        (ctype, mfr, model)
                    ).fetchone()
                cat_id = row['id'] if row else None

            if cat_id:
                cursor.execute(
                    "INSERT INTO base_config_components (config_id, component_id, quantity) VALUES (?, ?, ?)",
                    (config_id, cat_id, comp.get('quantity', 1))
                )
        db.commit()
        comp_count = len(data.get('components', []))
        # 5 base (add_config) + 1 per component (add_config_component)
        log_transaction(
            'add_config',
            config_id=config_id,
            metadata={
                'config_name': data['config_name'],
                'project_id': project_id,
                'component_count': comp_count,
            },
            tokens_override=5 + comp_count
        )
        return jsonify({'success': True, 'config_id': config_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/configs/<int:config_id>', methods=['DELETE'])
def api_delete_config(config_id):
    db = get_db()
    try:
        db.execute("DELETE FROM base_configs WHERE id = ?", (config_id,))
        db.commit()
        log_transaction('delete_config', config_id=config_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/configs/<int:config_id>', methods=['PUT'])
def api_update_config(config_id):
    """Replace a config's name and all its components."""
    data = request.get_json()
    if not data or not data.get('config_name'):
        return jsonify({'error': 'config_name required'}), 400
    db = get_db()
    try:
        cursor = db.cursor()
        existing = cursor.execute(
            "SELECT id FROM base_configs WHERE id = ?", (config_id,)
        ).fetchone()
        if not existing:
            return jsonify({'error': 'Config not found'}), 404

        cursor.execute("UPDATE base_configs SET config_name = ? WHERE id = ?",
                       (data['config_name'].strip(), config_id))
        cursor.execute("DELETE FROM base_config_components WHERE config_id = ?", (config_id,))

        for comp in data.get('components', []):
            ctype = comp.get('component_type')
            pn    = comp.get('part_number') or None
            model = comp.get('model') or pn or comp.get('specs') or 'unknown'
            specs = comp.get('specs') or None
            mfr   = comp.get('manufacturer') or comp.get('manufacturer_name') or None

            cursor.execute("""
                INSERT OR IGNORE INTO component_catalog
                    (component_type, manufacturer, model, part_number, description, data_source)
                VALUES (?, ?, ?, ?, ?, 'config')
            """, (ctype, mfr, model, pn, specs))

            if cursor.rowcount == 1:
                cat_id = cursor.lastrowid
            else:
                if mfr is None:
                    row = cursor.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer IS NULL AND model=?",
                        (ctype, model)
                    ).fetchone()
                else:
                    row = cursor.execute(
                        "SELECT id FROM component_catalog "
                        "WHERE component_type=? AND manufacturer=? AND model=?",
                        (ctype, mfr, model)
                    ).fetchone()
                cat_id = row['id'] if row else None

            if cat_id:
                cursor.execute(
                    "INSERT INTO base_config_components (config_id, component_id, quantity) "
                    "VALUES (?, ?, ?)",
                    (config_id, cat_id, max(1, int(comp.get('quantity', 1))))
                )

        db.commit()
        log_transaction('edit_config', config_id=config_id,
                        metadata={'config_name': data['config_name']})
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"api_update_config error: {e}")
        return jsonify({'error': 'Failed to update config'}), 500
    finally:
        db.close()


# =============================================================================
# CONFIG MATCHING — score vendor quotes against a desired base configuration
# =============================================================================

@app.route('/api/configs/<int:config_id>/match')
def api_config_match(config_id):
    """
    Score every quote in the same project against a base config.

    Matching logic (two levels):
    - Type coverage  : % of config component_types present as quote categories
    - Part coverage  : % of config part_numbers found in quote product_numbers (exact)

    Also logs a config_match transaction.
    """
    db = get_db()
    try:
        # Resolve config → project
        config_row = db.execute(
            "SELECT id, config_name, project_id FROM base_configs WHERE id = ?", (config_id,)
        ).fetchone()
        if not config_row:
            return jsonify({'error': 'Config not found'}), 404

        project_id = config_row['project_id']

        # Config components
        config_comps = db.execute("""
            SELECT cc.component_type, cc.part_number, cc.model,
                   cc.manufacturer AS manufacturer_name, bcc.quantity
            FROM base_config_components bcc
            JOIN component_catalog cc ON bcc.component_id = cc.id
            WHERE bcc.config_id = ?
            ORDER BY cc.component_type
        """, (config_id,)).fetchall()
        config_comps = [dict(c) for c in config_comps]

        if not config_comps:
            return jsonify({
                'config_id': config_id,
                'config_name': config_row['config_name'],
                'quotes': [],
                'message': 'Config has no components defined'
            })

        # All active quotes in this project
        quotes = db.execute("""
            SELECT id, quote_id, vendor, total_amount, currency,
                   quote_date, expiry_date, description
            FROM quotes
            WHERE project_id = ? AND status != 'archived'
            ORDER BY quote_date DESC
        """, (project_id,)).fetchall()

        results = []
        for quote in quotes:
            qid = quote['id']

            # Quote line items
            line_items = db.execute("""
                SELECT category, product_number, description, quantity, unit_price, total_price
                FROM line_items
                WHERE quote_id = ?
            """, (qid,)).fetchall()

            quote_categories  = {li['category'] for li in line_items if li['category']}
            quote_part_numbers = {
                (li['product_number'] or '').strip().upper()
                for li in line_items if li['product_number']
            }

            # Score each config component
            component_results = []
            type_hits = 0
            part_hits = 0
            for comp in config_comps:
                ctype = comp['component_type'] or ''
                cpn   = (comp['part_number'] or '').strip().upper()

                type_match = ctype in quote_categories
                part_match = bool(cpn) and cpn in quote_part_numbers

                if type_match:
                    type_hits += 1
                if part_match:
                    part_hits += 1

                # Find matching line items for context
                matched_items = [
                    dict(li) for li in line_items
                    if li['category'] == ctype or
                       (cpn and (li['product_number'] or '').strip().upper() == cpn)
                ]

                component_results.append({
                    'component_type':  ctype,
                    'part_number':     comp['part_number'],
                    'model':           comp['model'],
                    'manufacturer':    comp['manufacturer_name'],
                    'quantity_required': comp['quantity'],
                    'type_match':      type_match,
                    'part_match':      part_match,
                    'matched_items':   matched_items,
                })

            total = len(config_comps)
            type_coverage_pct = round(type_hits / total * 100, 1) if total else 0
            part_coverage_pct = round(part_hits / total * 100, 1) if total else 0

            results.append({
                'quote_id':         qid,
                'quote_ref':        quote['quote_id'],
                'vendor':           quote['vendor'],
                'total_amount':     quote['total_amount'],
                'currency':         quote['currency'],
                'quote_date':       quote['quote_date'],
                'expiry_date':      quote['expiry_date'],
                'description':      quote['description'],
                'type_coverage_pct': type_coverage_pct,
                'part_coverage_pct': part_coverage_pct,
                'type_hits':        type_hits,
                'part_hits':        part_hits,
                'total_components': total,
                'components':       component_results,
            })

        # Sort by type coverage descending
        results.sort(key=lambda r: r['type_coverage_pct'], reverse=True)

        log_transaction(
            'config_match',
            config_id=config_id,
            metadata={
                'config_name': config_row['config_name'],
                'project_id': project_id,
                'quotes_evaluated': len(results),
            }
        )

        return jsonify({
            'config_id':   config_id,
            'config_name': config_row['config_name'],
            'project_id':  project_id,
            'quotes':      results,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# =============================================================================
# PRICE HISTORY
# =============================================================================

@app.route('/api/configs/<int:config_id>/price-history')
def api_config_price_history(config_id):
    """
    Return price-over-time data for quotes directly assigned to this config.
    Response: { config_id, config_name, series: [{vendor, points: [{quote_date, unit_cost, total_amount, quote_items, quote_ref}]}] }
    """
    db = get_db()
    try:
        config_row = db.execute(
            "SELECT id, config_name FROM base_configs WHERE id = ?", (config_id,)
        ).fetchone()
        if not config_row:
            return jsonify({'error': 'Config not found'}), 404

        quotes = db.execute("""
            SELECT quote_id, vendor, total_amount, quote_items, quote_date
            FROM quotes
            WHERE config_id = ? AND status != 'archived' AND total_amount IS NOT NULL
            ORDER BY quote_date ASC
        """, (config_id,)).fetchall()

        vendors = {}
        for quote in quotes:
            total_amt = float(quote['total_amount'])
            items     = quote['quote_items']
            unit_cost = round(total_amt / items, 2) if items and items > 0 else total_amt
            vendor    = quote['vendor'] or 'Unknown'
            vendors.setdefault(vendor, []).append({
                'quote_ref':    quote['quote_id'],
                'unit_cost':    unit_cost,
                'total_amount': total_amt,
                'quote_items':  items,
                'quote_date':   quote['quote_date'],
            })

        series = sorted(
            [{'vendor': v, 'points': pts} for v, pts in vendors.items()],
            key=lambda s: s['vendor']
        )

        return jsonify({
            'config_id':   config_id,
            'config_name': config_row['config_name'],
            'series':      series,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# =============================================================================
# EXPIRY DASHBOARD
# =============================================================================

@app.route('/api/reports/expiry')
def api_expiry_dashboard():
    """
    Returns two lists:
      upcoming  — active quotes expiring in the next 90 days (soonest first)
      expired   — active quotes that expired in the last 120 days (most recent first)
    Both are annotated with days_remaining (negative = already expired).
    """
    db = get_db()
    try:
        base_select = """
            SELECT
                q.id, q.quote_id, q.vendor, q.customer_name,
                q.quote_date, q.expiry_date, q.total_amount,
                q.tenant_name, q.project_name, q.ica, q.description,
                CAST(julianday(q.expiry_date) - julianday('now') AS INTEGER) AS days_remaining
            FROM quotes q
            WHERE q.status != 'archived'
              AND q.expiry_date IS NOT NULL
              AND q.expiry_date != ''
        """
        upcoming = db.execute(base_select + """
              AND julianday(q.expiry_date) >= julianday('now')
              AND julianday(q.expiry_date) <= julianday('now', '+90 days')
            ORDER BY q.expiry_date ASC
        """).fetchall()

        expired = db.execute(base_select + """
              AND julianday(q.expiry_date) < julianday('now')
              AND julianday(q.expiry_date) >= julianday('now', '-120 days')
            ORDER BY q.expiry_date DESC
        """).fetchall()

        return jsonify({
            'upcoming': [dict(r) for r in upcoming],
            'expired':  [dict(r) for r in expired],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# =============================================================================
# VENDOR SCORECARD
# =============================================================================

@app.route('/api/scorecard')
def api_scorecard():
    """
    Per-vendor aggregate stats across all active quotes:
      - quote_count
      - avg_expiry_days   : average days between quote_date and expiry_date
      - win_rate_pct      : % of projects where vendor had the lowest price
      - avg_delta_pct     : average % above/below project-mean price
                            (negative = cheaper than average, positive = more expensive)
    Only projects with quotes from at least 2 distinct vendors contribute to
    win_rate and avg_delta so the comparison is meaningful.
    """
    db = get_db()
    try:
        # ── Base stats ────────────────────────────────────────────────────────
        base = db.execute("""
            SELECT
                vendor,
                COUNT(*)                                          AS quote_count,
                ROUND(AVG(CAST(total_amount AS REAL) / COALESCE(NULLIF(quote_items, 0), 1)), 2) AS avg_unit_value,
                ROUND(AVG(
                    CASE
                        WHEN expiry_date IS NOT NULL AND expiry_date != ''
                         AND quote_date  IS NOT NULL AND quote_date  != ''
                        THEN julianday(expiry_date) - julianday(quote_date)
                    END
                ), 0)                                             AS avg_expiry_days
            FROM quotes
            WHERE status != 'archived'
              AND vendor IS NOT NULL AND vendor != ''
            GROUP BY vendor
            ORDER BY quote_count DESC
        """).fetchall()

        # ── Win rate (lowest price per project, competitive projects only) ────
        win_rows = db.execute("""
            WITH competitive AS (
                -- projects where at least 2 vendors quoted
                SELECT project_id
                FROM quotes
                WHERE status != 'archived'
                  AND project_id IS NOT NULL
                  AND total_amount IS NOT NULL
                  AND vendor IS NOT NULL
                GROUP BY project_id
                HAVING COUNT(DISTINCT vendor) >= 2
            ),
            proj_min AS (
                SELECT q.project_id, MIN(q.total_amount) AS min_price
                FROM quotes q
                JOIN competitive c ON c.project_id = q.project_id
                WHERE q.status != 'archived' AND q.total_amount IS NOT NULL
                GROUP BY q.project_id
            ),
            wins AS (
                SELECT q.vendor, COUNT(*) AS win_count
                FROM quotes q
                JOIN proj_min pm
                  ON pm.project_id = q.project_id
                 AND pm.min_price  = q.total_amount
                WHERE q.status != 'archived' AND q.vendor IS NOT NULL
                GROUP BY q.vendor
            ),
            appearances AS (
                SELECT q.vendor, COUNT(DISTINCT q.project_id) AS proj_count
                FROM quotes q
                JOIN competitive c ON c.project_id = q.project_id
                WHERE q.status != 'archived' AND q.vendor IS NOT NULL
                GROUP BY q.vendor
            )
            SELECT
                a.vendor,
                COALESCE(w.win_count, 0)                                          AS wins,
                a.proj_count,
                ROUND(COALESCE(w.win_count, 0) * 100.0 / a.proj_count, 1)        AS win_rate_pct
            FROM appearances a
            LEFT JOIN wins w ON w.vendor = a.vendor
        """).fetchall()
        win_map = {r['vendor']: dict(r) for r in win_rows}

        # ── Price delta vs project average (competitive projects only) ─────────
        delta_rows = db.execute("""
            WITH competitive AS (
                SELECT project_id
                FROM quotes
                WHERE status != 'archived'
                  AND project_id IS NOT NULL
                  AND total_amount IS NOT NULL
                  AND vendor IS NOT NULL
                GROUP BY project_id
                HAVING COUNT(DISTINCT vendor) >= 2
            ),
            proj_avg AS (
                SELECT q.project_id, AVG(q.total_amount) AS avg_price
                FROM quotes q
                JOIN competitive c ON c.project_id = q.project_id
                WHERE q.status != 'archived' AND q.total_amount IS NOT NULL
                GROUP BY q.project_id
            )
            SELECT
                q.vendor,
                ROUND(AVG((q.total_amount - pa.avg_price) / pa.avg_price * 100.0), 1) AS avg_delta_pct
            FROM quotes q
            JOIN proj_avg pa ON pa.project_id = q.project_id
            WHERE q.status != 'archived'
              AND q.vendor IS NOT NULL
              AND q.total_amount IS NOT NULL
            GROUP BY q.vendor
        """).fetchall()
        delta_map = {r['vendor']: r['avg_delta_pct'] for r in delta_rows}

        result = []
        for row in base:
            v = dict(row)
            vendor = v['vendor']
            wd = win_map.get(vendor, {})
            v['wins']             = wd.get('wins', 0)
            v['projects_competed']= wd.get('proj_count', 0)
            v['win_rate_pct']     = wd.get('win_rate_pct')
            v['avg_delta_pct']    = delta_map.get(vendor)
            result.append(v)

        return jsonify(result)
    except Exception as e:
        logger.error(f"Scorecard error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


# =============================================================================
# ADMIN: TRANSACTIONS LEDGER
# =============================================================================

@app.route('/admin/transactions')
def admin_transactions():
    """Transaction ledger view."""
    db = get_db()
    rows = db.execute("""
        SELECT t.id, t.user_name, t.tokens_charged, t.created_at,
               t.quote_id, t.config_id, t.metadata_json,
               tt.code AS type_code, tt.label AS type_label, tt.token_cost
        FROM transactions t
        JOIN transaction_types tt ON tt.id = t.type_id
        ORDER BY t.created_at DESC
        LIMIT 500
    """).fetchall()

    summary = db.execute("""
        SELECT COUNT(*) AS total_txns, COALESCE(SUM(tokens_charged),0) AS total_tokens
        FROM transactions
    """).fetchone()

    type_summary = db.execute("""
        SELECT tt.label, tt.code, COUNT(*) AS cnt,
               COALESCE(SUM(t.tokens_charged),0) AS total_tokens
        FROM transactions t
        JOIN transaction_types tt ON tt.id = t.type_id
        GROUP BY tt.id
        ORDER BY total_tokens DESC
    """).fetchall()

    db.close()

    transactions = []
    for r in rows:
        d = dict(r)
        try:
            d['metadata'] = json.loads(r['metadata_json']) if r['metadata_json'] else {}
        except Exception:
            d['metadata'] = {}
        transactions.append(d)

    return render_template(
        'admin/transactions_list.html',
        transactions=transactions,
        summary=dict(summary),
        type_summary=[dict(ts) for ts in type_summary],
    )


@app.route('/api/admin/transaction-types')
def api_transaction_types():
    """List all transaction types with their token costs."""
    db = get_db()
    rows = db.execute(
        "SELECT id, code, label, description, token_cost FROM transaction_types ORDER BY token_cost DESC"
    ).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/admin/transactions/summary')
def api_transactions_summary():
    """Aggregated transaction stats."""
    db = get_db()
    overall = db.execute("""
        SELECT COUNT(*) AS total_txns, COALESCE(SUM(tokens_charged),0) AS total_tokens
        FROM transactions
    """).fetchone()
    by_type = db.execute("""
        SELECT tt.code, tt.label, COUNT(*) AS cnt,
               COALESCE(SUM(t.tokens_charged),0) AS total_tokens
        FROM transactions t
        JOIN transaction_types tt ON tt.id = t.type_id
        GROUP BY tt.id
        ORDER BY total_tokens DESC
    """).fetchall()
    db.close()
    return jsonify({
        'overall': dict(overall),
        'by_type': [dict(r) for r in by_type],
    })


# =============================================================================
# ADMIN: ENTERED COMPONENTS
# =============================================================================

@app.route('/admin/entered-components')
def admin_entered_components():
    db = get_db()
    rows = db.execute("""
        SELECT cc.id, cc.component_type, cc.part_number, cc.description AS specs, cc.model,
               cc.created_at, cc.manufacturer AS manufacturer_name,
               bc.config_name, p.name AS project_name, t.name AS tenant_name
        FROM component_catalog cc
        JOIN base_config_components bcc ON bcc.component_id = cc.id
        JOIN base_configs         bc  ON bcc.config_id = bc.id
        JOIN projects             p   ON bc.project_id = p.id
        JOIN tenants              t   ON p.tenant_id = t.id
        ORDER BY cc.created_at DESC
    """).fetchall()
    db.close()
    return render_template('admin/entered_components.html', components=[dict(r) for r in rows])


@app.route('/admin/components')
def admin_components():
    """List all components from component_catalog."""
    db = get_db()
    rows = db.execute("""
        SELECT cc.component_type, cc.manufacturer, cc.part_number, cc.model,
               cc.description AS specs, cc.data_source,
               COUNT(DISTINCT bcc.config_id) AS config_refs,
               COUNT(DISTINCT sqc.server_id) AS server_refs,
               GROUP_CONCAT(DISTINCT s.model_name) AS server_models
        FROM component_catalog cc
        LEFT JOIN base_config_components bcc ON bcc.component_id = cc.id
        LEFT JOIN server_quickspec_components sqc ON sqc.catalog_id = cc.id
        LEFT JOIN servers s ON s.id = sqc.server_id
        GROUP BY cc.id
        ORDER BY cc.component_type, cc.part_number
    """).fetchall()
    db.close()
    return render_template('admin/components_list.html', components=[dict(r) for r in rows])


@app.route('/admin/tenants/<int:tenant_id>/projects')
def admin_tenant_projects(tenant_id):
    """Show projects for a specific tenant with their quotes."""
    db = get_db()

    # Get tenant info
    tenant = db.execute('SELECT * FROM tenants WHERE id = ?', (tenant_id,)).fetchone()
    if not tenant:
        return "Tenant not found", 404

    # Get tenant's projects with quote counts
    projects = db.execute('''
        SELECT p.id, p.name, p.description, p.status, p.created_at,
               COUNT(q.id) as quote_count
        FROM projects p
        LEFT JOIN quotes q ON q.project_id = p.id
        WHERE p.tenant_id = ? AND p.status != 'archived'
        GROUP BY p.id
        ORDER BY p.name
    ''', (tenant_id,)).fetchall()

    # Get quotes for each project
    projects_with_quotes = []
    for project in projects:
        project_dict = dict(project)

        # Get quotes for this project
        quotes = db.execute('''
            SELECT id, quote_id, vendor, customer_name, quote_date, expiry_date,
                   total_amount, currency, description, ica, uploaded_at, po_comments
            FROM quotes
            WHERE project_id = ?
            ORDER BY uploaded_at DESC
        ''', (project['id'],)).fetchall()

        project_dict['quotes'] = [dict(q) for q in quotes]
        projects_with_quotes.append(project_dict)

    db.close()
    return render_template('admin/tenant_projects.html',
                          tenant=dict(tenant),
                          projects=projects_with_quotes)


@app.route('/admin/projects/<int:project_id>/quotes')
def admin_project_quotes(project_id):
    """Show quotes for a specific project."""
    db = get_db()

    # Get project info with tenant
    project = db.execute('''
        SELECT p.*, t.name as tenant_name
        FROM projects p
        LEFT JOIN tenants t ON t.id = p.tenant_id
        WHERE p.id = ?
    ''', (project_id,)).fetchone()

    if not project:
        return "Project not found", 404

    # Get project's quotes
    quotes = db.execute('''
        SELECT id, quote_id, vendor, customer_name, quote_date, expiry_date,
               total_amount, currency, description, ica, uploaded_at, po_comments
        FROM quotes
        WHERE project_id = ?
        ORDER BY quote_date DESC
    ''', (project_id,)).fetchall()

    db.close()
    return render_template('admin/project_quotes.html',
                          project=dict(project),
                          quotes=[dict(q) for q in quotes])


@app.route('/admin/projects/new')
def admin_project_new():
    """Show form to create new project."""
    db = get_db()
    tenants = db.execute("SELECT id, name FROM tenants WHERE status != 'archived' ORDER BY name").fetchall()
    db.close()
    return render_template('admin/project_form.html', tenants=[dict(t) for t in tenants])


@app.route('/admin/projects/create', methods=['POST'])
def admin_project_create():
    """Create a new project."""
    project_name = request.form.get('project_name', '').strip()
    tenant_id = request.form.get('tenant_id', '').strip()
    description = request.form.get('description', '').strip()
    delivery_deadline = request.form.get('delivery_deadline', '').strip() or None
    budget_raw = request.form.get('budget', '').strip()
    budget = float(budget_raw) if budget_raw else None

    if not project_name or not tenant_id:
        return jsonify({'error': 'Project name and tenant are required'}), 400

    db = get_db()
    try:
        cursor = db.cursor()
        cursor.execute('''
            INSERT INTO projects (name, tenant_id, description, delivery_deadline, budget, status)
            VALUES (?, ?, ?, ?, ?, 'active')
        ''', (project_name, int(tenant_id), description, delivery_deadline, budget))
        db.commit()
        project_id = cursor.lastrowid
        logger.info(f"Created project: {project_name} for tenant {tenant_id} (ID: {project_id})")
        log_transaction('add_project', metadata={
            'project_id': project_id, 'name': project_name, 'tenant_id': int(tenant_id)
        })
        return jsonify({'success': True, 'project_id': project_id})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Project with this name already exists for this tenant'}), 400
    except Exception as e:
        logger.error(f"Error creating project: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/admin/projects/<int:project_id>/edit')
def admin_project_edit(project_id):
    """Show edit form for a project."""
    db = get_db()
    project = db.execute(
        'SELECT id, name, comments, delivery_deadline, budget, status FROM projects WHERE id = ?', (project_id,)
    ).fetchone()
    db.close()
    if not project:
        return "Project not found", 404
    return render_template('admin/project_edit.html', project=dict(project))


@app.route('/admin/projects/<int:project_id>/update', methods=['POST'])
def admin_project_update(project_id):
    """Save project edits and redirect to admin."""
    project_name = request.form.get('project_name', '').strip()
    comments = request.form.get('project_comments', '').strip()
    delivery_deadline = request.form.get('delivery_deadline', '').strip() or None
    budget_raw = request.form.get('budget', '').strip()
    budget = float(budget_raw) if budget_raw else None

    if not project_name:
        return "Project name is required", 400

    db = get_db()
    try:
        db.execute('''
            UPDATE projects
            SET name = ?, comments = ?, delivery_deadline = ?, budget = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (project_name, comments or None, delivery_deadline, budget, project_id))
        db.commit()
        logger.info(f"Updated project {project_id}: name={project_name}")
    except sqlite3.IntegrityError:
        return "A project with that name already exists for this tenant", 400
    finally:
        db.close()

    return redirect(url_for('admin_dashboard'))


# =============================================================================
# QUICKSPEC PARSER — parse HPE QuickSpec PDFs into server + component catalog
# =============================================================================

from quickspec_parser import QuickSpecParser


def _save_quickspec_to_db(server_data: dict, components: list, pdf_path: str) -> dict:
    """
    Upsert server + components into the catalog in a single transaction.

    - servers:                  INSERT OR IGNORE on model_name (UNIQUE)
    - component_catalog:        INSERT OR IGNORE on (component_type, manufacturer, model)
                                Uses part_number as model field per design spec
    - server_quickspec_components: INSERT OR IGNORE on (server_id, catalog_id)

    Returns summary: {server_name, server_id, components_new, total}
    """
    db = get_db()
    try:
        db.execute("PRAGMA foreign_keys = ON")

        vendor = server_data.get('manufacturer', 'HPE')
        mfr_row = db.execute(
            "SELECT id FROM manufacturers WHERE name = ?", (vendor,)
        ).fetchone()
        if not mfr_row:
            db.execute("INSERT OR IGNORE INTO manufacturers (name) VALUES (?)", (vendor,))
            mfr_row = db.execute("SELECT id FROM manufacturers WHERE name = ?", (vendor,)).fetchone()
        manufacturer_id = mfr_row['id'] if mfr_row else None

        # Upsert server
        db.execute("""
            INSERT OR IGNORE INTO servers
                (manufacturer_id, model_name, model_number, form_factor, generation, pdf_path)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            manufacturer_id,
            server_data['model_name'],
            server_data.get('model_number'),
            server_data.get('form_factor'),
            server_data.get('generation'),
            pdf_path,
        ))

        server_row = db.execute(
            "SELECT id FROM servers WHERE model_name = ?", (server_data['model_name'],)
        ).fetchone()
        server_id = server_row['id']

        # Replace strategy: clear all existing component links for this server so a
        # re-upload fully replaces the classification (avoids stale Disk entries for NICs etc.)
        db.execute("DELETE FROM server_quickspec_components WHERE server_id = ?", (server_id,))

        TYPE_NORMALIZER = {
            'CPU': 'CPU', 'Memory': 'Memory', 'Disk': 'Disk',
            'Storage Controller': 'Storage Controller', 'Network Card': 'Network Card',
            'Power Supply': 'Power Supply', 'GPU': 'GPU',
            'Additional Hardware': 'Additional Hardware',
        }

        components_new = 0

        for comp in components:
            ctype = TYPE_NORMALIZER.get(comp['component_type'], 'Additional Hardware')
            pn    = comp['part_number']
            desc  = comp.get('description', '')

            # Upsert catalog entry — part_number used as model per design spec
            db.execute("""
                INSERT OR IGNORE INTO component_catalog
                    (component_type, manufacturer, model, part_number, description, data_source)
                VALUES (?, ?, ?, ?, ?, 'quickspec_pdf')
            """, (ctype, vendor, pn, pn, desc))

            cat_row = db.execute(
                "SELECT id FROM component_catalog "
                "WHERE component_type = ? AND manufacturer = ? AND model = ?",
                (ctype, vendor, pn)
            ).fetchone()
            if not cat_row:
                continue
            catalog_id = cat_row['id']

            db.execute("""
                INSERT OR IGNORE INTO server_quickspec_components
                    (server_id, catalog_id, component_role, is_standard, is_optional)
                VALUES (?, ?, ?, ?, ?)
            """, (
                server_id, catalog_id,
                comp.get('component_role'),
                1 if comp.get('is_standard') else 0,
                1 if comp.get('is_optional') else 0,
            ))

            components_new += 1

        db.commit()
        return {
            'server_name':         server_data['model_name'],
            'server_id':           server_id,
            'model_number':        server_data.get('model_number'),
            'form_factor':         server_data.get('form_factor'),
            'generation':          server_data.get('generation'),
            'components_new': components_new,
            'total':          len(components),
        }
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@app.route('/admin/quickspec/upload')
def admin_quickspec_upload_view():
    """QuickSpec upload form view (injected into admin dashboard)."""
    return render_template('admin/quickspec_upload.html')


@app.route('/api/admin/quickspec/upload', methods=['POST'])
def api_quickspec_upload():
    """
    Upload and parse an HPE QuickSpec PDF.
    Auto-saves on success. Returns JSON summary.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename or not allowed_file(file.filename):
        return jsonify({'error': 'Only PDF files allowed'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(DATA_DIR, 'quickspecs', filename)
    file.save(filepath)

    try:
        is_valid, err = validate_pdf(filepath)
        if not is_valid:
            os.remove(filepath)
            return jsonify({'error': f'PDF validation failed: {err}'}), 400

        parser = QuickSpecParser(filepath)
        result = parser.parse()

        server     = result['server']
        components = result['components']

        if not server.get('model_name'):
            os.remove(filepath)
            return jsonify({'error': 'Could not detect server model from PDF. '
                                     'Supports HPE QuickSpec and Dell Technical Guide PDFs.'}), 422

        summary = _save_quickspec_to_db(server, components, filepath)

        log_transaction(
            'parse_quickspec',
            metadata={
                'filename':        filename,
                'server_model':    server['model_name'],
                'components_new':  summary['components_new'],
                'total':           summary['total'],
                'page_count':      result.get('page_count'),
            },
            tokens_override=5 + summary['total']
        )

        logger.info(
            f"QuickSpec saved: {server['model_name']} — "
            f"{summary['components_new']} components saved"
        )
        return jsonify(summary)

    except Exception as e:
        logger.error(f"QuickSpec parse error for {filename}: {e}", exc_info=True)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass
        return jsonify({'error': str(e)}), 500


@app.route('/admin/servers')
def admin_servers():
    """Server catalog list view."""
    db = get_db()
    servers = db.execute("""
        SELECT s.id, s.model_name, s.model_number, s.form_factor, s.generation,
               s.created_at, s.pdf_path, m.name AS manufacturer_name,
               COUNT(sqc.id) AS component_count
        FROM servers s
        LEFT JOIN manufacturers m ON s.manufacturer_id = m.id
        LEFT JOIN server_quickspec_components sqc ON sqc.server_id = s.id
        GROUP BY s.id
        ORDER BY s.model_name
    """).fetchall()
    db.close()
    return render_template('admin/servers_list.html', servers=[dict(s) for s in servers])


@app.route('/api/catalog/components')
def api_catalog_components():
    """Return components of a given type that appear in at least one server spec.

    For Network Card, applies an additional description filter so that stale
    catalog entries (disks, bezel kits, etc. incorrectly stored as Network Card
    from old parses) don't pollute the NIC picker dropdown.
    """
    ctype = request.args.get('type', '')
    db = get_db()
    rows = db.execute("""
        SELECT DISTINCT cc.part_number, cc.description, cc.manufacturer
        FROM component_catalog cc
        JOIN server_quickspec_components sqc ON sqc.catalog_id = cc.id
        WHERE cc.component_type = ?
        ORDER BY cc.manufacturer, cc.part_number
    """, (ctype,)).fetchall()
    db.close()

    result = [dict(r) for r in rows]

    if ctype == 'Network Card':
        import re
        _nic_re = re.compile(
            r'ethernet|infiniband|adapter|omni-path|\bnic\b|gbe\b|10g|25g|40g|100g|'
            r'sfp|qsfp|osfp|ocp\s*\d|\bfcoe\b|\broce\b',
            re.IGNORECASE
        )
        result = [r for r in result if r['description'] and _nic_re.search(r['description'])]
        # Deduplicate by part_number — prefer named manufacturer over 'Unknown'
        seen = {}
        for r in result:
            pn = r['part_number']
            if pn not in seen or seen[pn]['manufacturer'] in (None, 'Unknown'):
                seen[pn] = r
        result = sorted(seen.values(), key=lambda r: (r['manufacturer'] or '', r['part_number']))

    return jsonify(result)


@app.route('/api/admin/servers')
def api_admin_servers():
    """JSON list of servers with component counts. Accepts ?q= for search."""
    q      = request.args.get('q', '').strip()
    db     = get_db()
    sql    = """
        SELECT s.id, s.model_name, s.model_number, s.form_factor, s.generation,
               s.created_at, m.name AS manufacturer_name,
               COUNT(sqc.id) AS component_count
        FROM servers s
        LEFT JOIN manufacturers m ON s.manufacturer_id = m.id
        LEFT JOIN server_quickspec_components sqc ON sqc.server_id = s.id
    """
    params = []
    if q:
        sql += " WHERE s.model_name LIKE ? OR s.model_number LIKE ? OR m.name LIKE ?"
        like = f'%{q}%'
        params.extend([like, like, like])
    sql += " GROUP BY s.id ORDER BY s.model_name"
    if q:
        sql += " LIMIT 10"
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/servers/<int:server_id>/components')
def api_server_components_json(server_id):
    """JSON: components for a server grouped by component_type."""
    db = get_db()
    rows = db.execute("""
        SELECT cc.component_type, cc.part_number AS part_number,
               cc.description, cc.manufacturer,
               sqc.component_role, sqc.is_standard, sqc.is_optional
        FROM server_quickspec_components sqc
        JOIN component_catalog cc ON sqc.catalog_id = cc.id
        WHERE sqc.server_id = ?
        ORDER BY cc.component_type,
                 sqc.is_standard DESC,
                 cc.part_number
    """, (server_id,)).fetchall()
    db.close()
    from collections import defaultdict
    grouped = defaultdict(list)
    for r in rows:
        grouped[r['component_type']].append(dict(r))
    return jsonify(grouped)


@app.route('/admin/servers/<int:server_id>/components')
def admin_server_components(server_id):
    """Detail view: all components linked to a server via QuickSpec."""
    db = get_db()
    server = db.execute("""
        SELECT s.*, m.name AS manufacturer_name
        FROM servers s
        LEFT JOIN manufacturers m ON s.manufacturer_id = m.id
        WHERE s.id = ?
    """, (server_id,)).fetchone()
    if not server:
        return 'Server not found', 404

    components = db.execute("""
        SELECT cc.component_type, cc.part_number, cc.description,
               sqc.component_role, sqc.is_standard, sqc.is_optional, sqc.created_at
        FROM server_quickspec_components sqc
        JOIN component_catalog cc ON sqc.catalog_id = cc.id
        WHERE sqc.server_id = ?
        ORDER BY cc.component_type, cc.part_number
    """, (server_id,)).fetchall()
    db.close()
    return render_template(
        'admin/server_components.html',
        server=dict(server),
        components=[dict(c) for c in components],
    )


@app.route('/admin/servers/<int:server_id>/pdf')
def admin_server_pdf(server_id):
    """Serve the QuickSpec PDF stored for a server."""
    from flask import send_file
    db = get_db()
    row = db.execute('SELECT pdf_path, model_name FROM servers WHERE id = ?', (server_id,)).fetchone()
    db.close()
    if not row or not row['pdf_path']:
        return 'PDF not found', 404
    pdf_path = row['pdf_path']
    if not os.path.isabs(pdf_path):
        pdf_path = os.path.join(_app_dir, pdf_path)
    if not os.path.exists(pdf_path):
        fallback = os.path.join(DATA_DIR, 'quickspecs', os.path.basename(pdf_path))
        if os.path.exists(fallback):
            pdf_path = fallback
        else:
            return 'PDF file not found on disk', 404
    return send_file(pdf_path, mimetype='application/pdf',
                     download_name=os.path.basename(pdf_path))


if __name__ == '__main__':
    # Initialize database if it doesn't exist, then apply migrations either way
    if not os.path.exists(app.config['DATABASE']):
        init_db()
    migrate_db()

    logger.info(f"Loaded DB: {app.config['DATABASE']}")
    app.run(debug=True, host='0.0.0.0', port=5001)
